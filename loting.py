# Main file for the program
import sqlite3 as sql
import xlsxwriter
from xlsxwriter.utility import xl_range
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from random import randint
import random
# TODO is random random?
import time
from tabulate import tabulate
from tkinter import *
from tkinter import messagebox
from classes.entrybox import EntryBox
import os.path
import configparser
import textwrap
import shutil
import pathlib
import re

template_password = 'NTDS2018'

# TODO Remove leading and trailing spaces from names and other options

# Program constants
total_width = 190
status_text_width = 140
default_db_name = 'NTDS'
db_ext = '.db'
xlsx_ext = '.xlsx'
ini_ext = '.ini'
database_key = {'db': '', 'path': '', 'session_timestamp': ''}
max_row = 201
runs = 1000
backup_ext = '_Backup' + xlsx_ext

# Formatting names
lions = 'Lions'
contestants = 'contestants'
individuals = 'individuals'

# Folder names
config_folder = 'config'
settings_folder = 'settings'
output_folder = 'output'
statistics_folder = 'statistics'
sheets_folder = 'signupsheets_captains'
used_sheets_folder = 'used_signupsheets'

# Setup configuration and setting files dictionaries
config_key = {'name': 'config.ini', 'folder': config_folder, 'path': ''}
config_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), config_key['folder'])
config_key['path'] = os.path.join(config_key['folder'], config_key['name'])

settings_key = {'name': 'user_settings.ini', 'folder': settings_folder, 'path': ''}
settings_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), settings_key['folder'])
settings_key['path'] = os.path.join(settings_key['folder'], settings_key['name'])

participating_teams_key = {'name': 'participating_teams.ini', 'folder': config_folder, 'path': ''}
participating_teams_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                                 participating_teams_key['folder'])
participating_teams_key['path'] = os.path.join(participating_teams_key['folder'], participating_teams_key['name'])

template_key = {'name': 'template.ini', 'folder': config_folder, 'path': ''}
template_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), template_key['folder'])
template_key['path'] = os.path.join(template_key['folder'], template_key['name'])

excel_template_key = {'name': 'NTDS_Template.xlsx', 'folder': config_folder, 'path': ''}
excel_template_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), excel_template_key['folder'])
excel_template_key['path'] = os.path.join(excel_template_key['folder'], excel_template_key['name'])

sheets_key = {'name': '', 'folder': sheets_folder, 'path': ''}
sheets_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), sheets_key['folder'])
sheets_key['path'] = os.path.join(sheets_key['folder'], sheets_key['name'])

output_key = {'name': 'NTDS_Selection', 'folder': output_folder, 'path': ''}
output_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_key['folder'])
output_key['path'] = os.path.join(output_key['folder'], output_key['name'])

statistics_key = {'name': 'NTDS_Statistics', 'folder': statistics_folder, 'path': ''}
statistics_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), statistics_key['folder'])
statistics_key['path'] = os.path.join(statistics_key['folder'], statistics_key['name'])

used_sheets_key = {'name': '', 'folder': used_sheets_folder, 'path': ''}
used_sheets_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), used_sheets_key['folder'])
used_sheets_key['path'] = os.path.join(used_sheets_key['folder'], used_sheets_key['name'])

status_key = {'name': 'NTDS_Status', 'folder': statistics_folder, 'path': ''}
status_key['folder'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), status_key['folder'])
status_key['path'] = os.path.join(status_key['folder'], status_key['name'])

# Create folders
pathlib.Path(config_key['folder']).mkdir(parents=True, exist_ok=True)
pathlib.Path(settings_key['folder']).mkdir(parents=True, exist_ok=True)
pathlib.Path(sheets_key['folder']).mkdir(parents=True, exist_ok=True)
pathlib.Path(output_key['folder']).mkdir(parents=True, exist_ok=True)
pathlib.Path(statistics_key['folder']).mkdir(parents=True, exist_ok=True)
pathlib.Path(used_sheets_key['folder']).mkdir(parents=True, exist_ok=True)

# WISHLIST
# TODO refactor (remove hardcoded material that snuck in)

# Create files if they are not there
if not os.path.isfile(path=config_key['path']):
    open(config_key['path'], 'w+').close()
if not os.path.isfile(path=settings_key['path']):
    open(settings_key['path'], 'w+').close()
if not os.path.isfile(path=participating_teams_key['path']):
    open(participating_teams_key['path'], 'w+').close()
if not os.path.isfile(path=template_key['path']):
    open(template_key['path'], 'w+').close()

# Participating teams
participating_teams_dict =\
    {'Team01': {'team_name': '4 happy feet', 'city': 'Enschede', 'signup_sheet': 'NTDS_Enschede.xlsx'},
     'Team02': {'team_name': 'AmsterDance', 'city': 'Amsterdam', 'signup_sheet': 'NTDS_Amsterdam.xlsx'},
     'Team03': {'team_name': 'Blue Suede Shoes', 'city': 'Delft', 'signup_sheet': 'NTDS_Delft.xlsx'},
     'Team04': {'team_name': 'Dance Fever', 'city': 'Nijmegen', 'signup_sheet': 'NTDS_Nijmegen.xlsx'},
     'Team05': {'team_name': 'Erasmus Dance Society', 'city': 'Rotterdam', 'signup_sheet': 'NTDS_Rotterdam.xlsx'},
     'Team06': {'team_name': 'Footloose', 'city': 'Eindhoven', 'signup_sheet': 'NTDS_Eindhoven.xlsx'},
     'Team07': {'team_name': 'LeiDance', 'city': 'Leiden', 'signup_sheet': 'NTDS_Leiden.xlsx'},
     'Team08': {'team_name': 'Let`s Dance', 'city': 'Maastricht', 'signup_sheet': 'NTDS_Maastricht.xlsx'},
     'Team09': {'team_name': 'The Blue Toes', 'city': 'Groningen', 'signup_sheet': 'NTDS_Groningen.xlsx'},
     'Team10': {'team_name': 'U Dance', 'city': 'Utrecht', 'signup_sheet': 'NTDS_Utrecht.xlsx'},
     'Team11': {'team_name': 'WUBDA', 'city': 'Wageningen', 'signup_sheet': 'NTDS_Wageningen.xlsx'},
     }

# Levels
levels = {'beginners': 'Beginners', 'breitensport': 'Breitensport', 'closed': 'CloseD', 'open_class': 'Open Class',
          0: ''}
# Roles
roles = {'lead': 'Lead', 'follow': 'Follow', 0: ''}
# Signup options
options_ymn = {'yes': 'Ja', 'maybe': 'Misschien', 'no': 'Nee'}
options_yn = {'yes': options_ymn['yes'], 'no': options_ymn['no']}

# Boundaries
boundaries = {'max_contestants': 300, 'min_guaranteed_beginners': 4, 'min_fixed_lion_contestants': 10,
              'beginner_signup_cutoff_percentage': 20, 'buffer_for_selection_percentage': 10}
# Participating classes
classes = [levels['beginners'], levels['breitensport'], levels['open_class']]
# Participants Lion points
lion_participants = [levels['beginners'], levels['breitensport']]

lvls = 'DancingClasses'
rls = 'DancingRoles'
ymn = 'YesMaybeNo'
yn = 'YesNo'
an = 'Any'
num = 'Number'

cont_num = 'ContestantNumbers'
avail_cls = 'AvailableClasses'
lio = 'Lions'

template = 'Template'


def read_add_config(parser, section, var, output=None, user_dict=None, var_type=None):
    """"Temp"""
    if output is None:
        if section in parser:
            var = dict(parser.items(section))
            if var_type is not None:
                for k, v in var.items():
                    try:
                        var[k] = var_type(v)
                    except KeyError:
                        pass
        else:
            parser.add_section(section)
            for k, v in var.items():
                parser.set(section, str(k), str(v))
    elif output == 'list' and user_dict is not None:
        if section in parser:
            var = list()
            for it in parser[section]:
                if parser[section].getboolean(it):
                    var.append(user_dict[it])
        else:
            parser.add_section(section)
            for ind, it in enumerate(list(user_dict.values())):
                if it in var:
                    parser.set(section, str(list(user_dict.keys())[ind]), str(True))
    return var


def add_config(parser, section, var):
    """"Temp"""
    parser.add_section(section)
    for k, v in var.items():
        parser.set(section, str(k), str(v))


# Create participating teams ini
config_parser = configparser.ConfigParser()
if os.path.isfile(path=participating_teams_key['path']):
    config_parser.read(participating_teams_key['path'])
    for key, value in participating_teams_dict.items():
        if key in config_parser:
            participating_teams_dict[key] = dict(config_parser.items(key))
        else:
            config_parser.add_section(key)
            for key2, value2 in participating_teams_dict[key].items():
                config_parser.set(key, str(key2), str(value2))
    with open(participating_teams_key['path'], 'w') as configfile:
        config_parser.write(configfile)

# Create config ini
config_parser = configparser.ConfigParser()
if os.path.isfile(path=config_key['path']):
    config_parser.read(config_key['path'])
    levels = read_add_config(config_parser, lvls, levels)
    roles = read_add_config(config_parser, rls, roles)
    options_ymn = read_add_config(config_parser, ymn, options_ymn)
    options_yn = read_add_config(config_parser, yn, options_yn)
    with open(config_key['path'], 'w') as configfile:
        config_parser.write(configfile)

# Create settings ini
config_parser = configparser.ConfigParser()
if os.path.isfile(path=settings_key['path']):
    config_parser.read(settings_key['path'])
    boundaries = read_add_config(config_parser, cont_num, boundaries, var_type=int)
    classes = read_add_config(config_parser, avail_cls, classes, output='list', user_dict=levels)
    lion_participants = read_add_config(config_parser, lio, lion_participants, output='list', user_dict=levels)
    with open(settings_key['path'], 'w') as configfile:
        config_parser.write(configfile)
boundaries['beginner_signup_cutoff'] = int(boundaries['max_contestants'] *
                                           boundaries['beginner_signup_cutoff_percentage'] / 100)
boundaries['buffer_for_selection'] = int(boundaries['max_contestants'] *
                                         boundaries['buffer_for_selection_percentage'] / 100)

# Title for Excel export files
template_dict = {'id': '0,Nr.,INT PRIMARY KEY,3.57,' + num,
                 'first_name': '1,Voornaam,TEXT,17.86,' + an,
                 'ln_prefix': '2,Tussenvoegsel,TEXT,13.57,' + an,
                 'last_name': '3,Achternaam,TEXT,17.86,' + an,
                 'email': '4,E-mail,TEXT,19.29,' + an,
                 'ballroom_level': '5,Welk Ballroom klassement dans je?,TEXT,12.14,' + lvls,
                 'latin_level': '6,Welk Latin klassement dans je?,TEXT,12.14,' + lvls,
                 'ballroom_partner': '7,Wat is het Nr. van je Ballroom partner?,INT,8.43,' + num,
                 'latin_partner': '8,Wat is het Nr. van je Latin partner?,INT,8.43,' + num,
                 'ballroom_role': '9,Wat is jouw Ballroom rol?,TEXT,8.43,' + rls,
                 'latin_role': '10,Wat is jouw Latin rol?,TEXT,8.43,' + rls,
                 'ballroom_mandatory_blind_date': '11,Moet je in de Ballroom verplicht blind daten?,TEXT,8.43,' + yn,
                 'latin_mandatory_blind_date': '12,Moet je in de Latin verplicht blind daten?,TEXT,8.43,' + yn,
                 'team_captain': '13,Ben je een Teamcaptain?,TEXT,13.00,' + yn,
                 'current_volunteer': '14,Wil je vrijwilliger zijn voor dit NTDS?,TEXT,10.71,' + ymn,
                 'past_volunteer': '15,Ben je op een eerder NTDS of ETDS vrijwilliger geweest?,TEXT,10.71,' + yn,
                 'first_aid': '16,Kun en wil je als EHBO vrijwilliger zijn?,TEXT,10.71,' + ymn,
                 'emergency_response_officer': '17,Kun en wil je als BHV vrijwilliger zijn?,TEXT,10.71,' + ymn,
                 'ballroom_jury': '18,Zou je vrijwilliger willen zijn als Ballroom jury?,TEXT,11.29,' + ymn,
                 'latin_jury': '19,Zou je vrijwilliger willen zijn als Latin jury?,TEXT,11.29,' + ymn,
                 'student': '20,Ben je student?,TEXT,8.43,' + yn,
                 'sleeping_location': '21,Blijf je slapen in een slaapzaal op het NTDS?,TEXT,8.43,' + yn,
                 'diet_wishes': '22,Allergiën / Dieet,TEXT,8.43,' + an,
                 'city': '23,City,TEXT,0,' + an}
config_parser = configparser.ConfigParser()
if os.path.isfile(path=template_key['path']):
    config_parser.read(template_key['path'])
    template_dict = read_add_config(config_parser, template, template_dict)
    with open(template_key['path'], 'w') as configfile:
        config_parser.write(configfile)

# Template values
template_values = [x.split(',')[1] for x in list(template_dict.values())]
template_values = template_values[:-1]
column_widths = [float(x.split(',')[3]) for x in list(template_dict.values())]
column_widths = column_widths[:-1]
column_options = [x.split(',')[4] for x in list(template_dict.values())]
column_options = column_options[:-1]

# Data validation values for Excel
validation_dict = {lvls: {'validate': 'list', 'source': classes},
                   rls: {'validate': 'list', 'source': list(roles.values())},
                   ymn: {'validate': 'list', 'source': list(options_ymn.values())},
                   yn: {'validate': 'list', 'source': list(options_yn.values())},
                   an: {'validate': 'any'},
                   num: {'validate': 'integer', 'criteria': '>', 'value': 0}}

# Create Excel template file
with xlsxwriter.Workbook(excel_template_key['path']) as wb:
    ws = wb.add_worksheet(name='Input')
    unlocked = wb.add_format({'locked': False})
    locked = wb.add_format({'locked': True})
    f = wb.add_format({'text_wrap': True})
    ws.set_column(xl_range(0, len(template_values), 0, 16383), None, None, {'hidden': True})
    ws.set_default_row(hide_unused_rows=True)
    ws.set_row(0, 90, locked)
    for c in range(0, len(template_values)):
        ws.write(0, c, template_values[c], f)
        ws.set_column(c, c, column_widths[c], unlocked)
    for r in range(1, max_row):
        ws.set_row(r, 15)
        ws.write_formula(r, 0, '=IF(B' + str(r+1) + '<>"",ROW(A' + str(r+1) + ')-1,"")')
    for c in range(0, len(template_values)):
        ws.data_validation(1, c, max_row, c, dict(validation_dict[column_options[c]]))
    ws.freeze_panes(1, 0)
    ws.set_column('A:A', column_widths[0], locked)
    ws.protect(template_password)
# Create signup sheets for team captains to fill in
for key, value in participating_teams_dict.items():
    shutil.copy2(excel_template_key['path'], os.path.join(sheets_key['folder'], value['signup_sheet']))
    shutil.copy2(excel_template_key['path'], os.path.join(sheets_key['folder'], value['signup_sheet'].replace(xlsx_ext,backup_ext)))

# General dictionary
gen_dict = dict()
for key, value in template_dict.items():
    gen_dict[key] = int(value.split(',')[0])

# SQL dictionary
sql_dict = dict()
for key, value in template_dict.items():
    sql_dict[key] = key

# Dictionary to create SQL query
dancers_list_dict = dict()
for key, value in template_dict.items():
    dancers_list_dict[key] = value.split(',')[2]

# Table names
signup_list = 'signup_list'
cancelled_list = 'cancelled_list'
selection_list = 'selection_list'
selected_list = 'selected_list'
backup_list = 'backup_list'
team_list = 'team_list'
partners_list = 'partners_list'
ref_partner_list = 'reference_partner_list'
fixed_beginners_list = 'fixed_beginners'
fixed_lions_list = 'fixed_lions'
beginners_list = 'beginners'
lions_list = 'lions'
contestants_list = 'contestants'
individual_list = 'individuals'

# Selectable options
options_dict = {lvls: levels, rls: roles, yn: options_yn, ymn: options_ymn}
NTDS_options = dict()
for key, value in template_dict.items():
    if value.split(',')[4] in [lvls, rls, yn, ymn]:
        NTDS_options[key] = options_dict[value.split(',')[4]]

# SQL Table column names and dictionary of teams list
team_dict = {'team': 0, 'city': 1, 'signup_list': 2}

# SQL Table column names and dictionary for partners list
partner_dict = {'num': 0, 'lead': 1, 'follow': 2, 'city_lead': 3, 'city_follow': 4,
                'ballroom_level_lead': 5, 'ballroom_level_follow': 6, 'latin_level_lead': 7, 'latin_level_follow': 8}
partner_sql_dict = {key: key for key in partner_dict}

# SQL Table column names and dictionary for city list
city_dict = {sql_dict['city']: 0, 'number_of_contestants': 1, 'max_contestants': 2}
city_sql_dict = {key: key for key in city_dict}

# SQL Table column names for individuals list
sql_run = 'run_number'

# General query formats
drop_table_query = 'DROP TABLE IF EXISTS {};'
dancers_list_query = 'CREATE TABLE {} ('
for key, value in dancers_list_dict.items():
    dancers_list_query += key + ' ' + value + ', '
dancers_list_query = dancers_list_query[:-2] + ');'
team_list_query = 'CREATE TABLE {tn} ({team} TEXT PRIMARY KEY, {city} TEXT, {signup_list} TEXT);' \
    .format(tn={}, team=list(team_dict.keys())[0],
            city=list(team_dict.keys())[1], signup_list=list(team_dict.keys())[2])
paren_table_query = 'CREATE TABLE {tn} ({no} INTEGER PRIMARY KEY AUTOINCREMENT, ' \
                    '{lead} INT, {follow} INT, {lead_city} TEXT, {follow_city} TEXT, ' \
                    '{ballroom_level_lead} TEXT, {ballroom_level_follow} TEXT, ' \
                    '{latin_level_lead} TEXT, {latin_level_follow} TEXT);'\
    .format(tn={}, no=partner_sql_dict['num'], lead=partner_sql_dict['lead'], follow=partner_sql_dict['follow'],
            lead_city=partner_sql_dict['city_lead'], follow_city=partner_sql_dict['city_follow'],
            ballroom_level_lead=partner_sql_dict['ballroom_level_lead'],
            ballroom_level_follow=partner_sql_dict['ballroom_level_follow'],
            latin_level_lead=partner_sql_dict['latin_level_lead'],
            latin_level_follow=partner_sql_dict['latin_level_follow'])
city_list_query = 'CREATE TABLE {tn} ({city} TEXT, {num} INT, {max_num} INT);' \
    .format(tn={}, city=sql_dict['city'], num=city_sql_dict['number_of_contestants'],
            max_num=city_sql_dict['max_contestants'])


status_dict = dict()
max_col = len(template_values)


def move_used_signupsheet(file_name):
    """"Temp"""
    new_path = os.path.join(used_sheets_key['folder'], time.strftime('%Y_%d_%m_%H_%M_%S', time.localtime(database_key['session_timestamp'])))
    if not os.path.exists(new_path):
        os.makedirs(new_path)
    os.rename(os.path.join(os.path.dirname(os.path.realpath(__file__)), file_name), os.path.join(new_path, file_name))


def find_partner(identifier, connection, cursor, city=None, signed_partner_only=False):
    """Finds the/a partner for a dancer, given id"""
    partner_id = None
    if identifier == 37:
        status_print('pause')
    status_print('')
    status_print('Looking for a partner for dancer {id}'.format(id=identifier))
    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selection_list, id=sql_dict['id'])
    dancer = cursor.execute(query, (identifier,)).fetchone()
    if dancer is not None:
        team = dancer[gen_dict['city']]
        ballroom_level = dancer[gen_dict['ballroom_level']]
        latin_level = dancer[gen_dict['latin_level']]
        if ballroom_level == '':
            ballroom_role = ''
        else:
            ballroom_role = dancer[gen_dict['ballroom_role']]
        if latin_level == '':
            latin_role = ''
        else:
            latin_role = dancer[gen_dict['latin_role']]
        ballroom_partner = dancer[gen_dict['ballroom_partner']]
        latin_partner = dancer[gen_dict['latin_partner']]
        # Check if the contestant's partner(s) are already selected for the tournament
        query = 'SELECT * FROM {tn} WHERE {id} = ? OR {id} = ?'.format(tn=selected_list, id=sql_dict['id'])
        partners_selected = cursor.execute(query, (ballroom_partner, latin_partner)).fetchall()
        if len(partners_selected) > 0:
            partner_id = identifier
        # Check if the contestant already has signed up with a partner (or two)
        if isinstance(ballroom_partner, int) and ballroom_partner == latin_partner and partner_id is None:
            partner_id = ballroom_partner
        if isinstance(ballroom_partner, int) and latin_partner == '' and partner_id is None:
            partner_id = ballroom_partner
            query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selection_list, id=sql_dict['id'])
            partner = cursor.execute(query, (partner_id,)).fetchone()
            partner_latin_partner = partner[gen_dict['latin_partner']]
            if isinstance(partner_latin_partner, int):
                status_print('{id1} and {id2} signed up together'
                             .format(id1=ballroom_partner, id2=partner_latin_partner))
                create_pair(ballroom_partner, partner_latin_partner, connection=connection, cursor=cursor)
                move_selected_contestant(partner_latin_partner, connection=connection, cursor=cursor)
        if ballroom_partner == '' and isinstance(latin_partner, int) and partner_id is None:
            partner_id = latin_partner
            query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selection_list, id=sql_dict['id'])
            partner = cursor.execute(query, (partner_id,)).fetchone()
            partner_ballroom_partner = partner[gen_dict['ballroom_partner']]
            if isinstance(partner_ballroom_partner, int):
                status_print('{id1} and {id2} signed up together'
                             .format(id1=latin_partner, id2=partner_ballroom_partner))
                create_pair(latin_partner, partner_ballroom_partner, connection=connection, cursor=cursor)
                move_selected_contestant(partner_ballroom_partner, connection=connection, cursor=cursor)
        if all([isinstance(ballroom_partner, int), isinstance(latin_partner, int), ballroom_partner != latin_partner,
                partner_id is None]):
            status_print('{id1} and {id2} signed up together'.format(id1=identifier, id2=ballroom_partner))
            create_pair(identifier, ballroom_partner, connection=connection, cursor=cursor)
            move_selected_contestant(ballroom_partner, connection=connection, cursor=cursor)
            partner_id = latin_partner
        if partner_id is not None:
            status_print('{id1} and {id2} signed up together'.format(id1=identifier, id2=partner_id))
        if partner_id is None:
            query = 'SELECT * FROM {tn} WHERE {ballroom_level} = ? AND {latin_level} = ? AND {ballroom_partner} = "" ' \
                    'AND {latin_partner} = "" ' \
                .format(tn=selection_list, ballroom_level=sql_dict['ballroom_level'],
                        latin_level=sql_dict['latin_level'],
                        ballroom_partner=sql_dict['ballroom_partner'], latin_partner=sql_dict['latin_partner'])
            if ballroom_role != '' and ballroom_role == latin_role:
                query += 'AND {ballroom_role} != ? AND latin_role != ? '\
                    .format(ballroom_role=sql_dict['ballroom_role'], latin_role=sql_dict['latin_role'])
            elif ballroom_role != '' and latin_role == '':
                query += 'AND {ballroom_role} != ? AND latin_role = ? ' \
                    .format(ballroom_role=sql_dict['ballroom_role'], latin_role=sql_dict['latin_role'])
            elif ballroom_role == '' and latin_role != '':
                query += 'AND {ballroom_role} = ? AND latin_role != ? ' \
                    .format(ballroom_role=sql_dict['ballroom_role'], latin_role=sql_dict['latin_role'])
            elif ballroom_role == '' and latin_role == '':
                query += 'AND {ballroom_role} = ? AND latin_role = ? ' \
                    .format(ballroom_role=sql_dict['ballroom_role'], latin_role=sql_dict['latin_role'])
            else:
                query += 'AND {ballroom_role} = ? AND latin_role != ? ' \
                    .format(ballroom_role=sql_dict['ballroom_role'], latin_role=sql_dict['latin_role'])
                query2 = 'SELECT * FROM {tn} WHERE {ballroom_level} = ? AND {latin_level} = ? AND ' \
                         '{ballroom_partner} = "" AND {latin_partner} = "" ' \
                         'AND {ballroom_role} = ? AND latin_role = ? '\
                    .format(tn=selection_list,
                            ballroom_level=sql_dict['ballroom_level'], latin_level=sql_dict['latin_level'],
                            ballroom_partner=sql_dict['ballroom_partner'], latin_partner=sql_dict['latin_partner'],
                            ballroom_role=sql_dict['ballroom_role'], latin_role=sql_dict['latin_role'])
                if city is None:
                    query2 += ' AND {team} != ?'.format(team=sql_dict['city'])
                else:
                    query2 += ' AND {team} = ?'.format(team=sql_dict['city'])
                    team = city
                potential_partners = cursor.execute(query2, (ballroom_level, '', ballroom_role, '', team)).fetchall()
                number_of_potential_partners = len(potential_partners)
                if number_of_potential_partners > 0:
                    random_num = randint(0, number_of_potential_partners - 1)
                    first_partner_id = potential_partners[random_num][gen_dict['id']]
                    if first_partner_id is not None:
                        status_print('Different roles: {id1} and {id2} matched together'
                                     .format(id1=identifier, id2=first_partner_id))
                        create_pair(identifier, first_partner_id, connection=connection, cursor=cursor)
                        move_selected_contestant(first_partner_id, connection=connection, cursor=cursor)
                ballroom_level = ''
                ballroom_role = ''
            if city is None:
                query += ' AND {team} != ?'.format(team=sql_dict['city'])
            else:
                query += ' AND {team} = ?'.format(team=sql_dict['city'])
                team = city
        # Try to find an "ideal" partner with the same combination of levels for the dancer
        potential_partners = []
        number_of_potential_partners = len(potential_partners)
        if signed_partner_only is False:
            if partner_id is None:
                potential_partners = cursor.\
                    execute(query, (ballroom_level, latin_level, ballroom_role, latin_role, team))\
                    .fetchall()
                number_of_potential_partners = len(potential_partners)
                if number_of_potential_partners > 0:
                    random_num = randint(0, number_of_potential_partners - 1)
                    partner_id = potential_partners[random_num][gen_dict['id']]
            if partner_id is None:
                query = 'SELECT * FROM {tn} WHERE {ballroom_level} = ? AND {latin_level} = ? ' \
                        'AND {ballroom_partner} = "" AND {latin_partner} = "" ' \
                        'AND {ballroom_role} != ? AND latin_role != ? '\
                    .format(tn=selection_list,
                            ballroom_level=sql_dict['ballroom_level'], latin_level=sql_dict['latin_level'],
                            ballroom_partner=sql_dict['ballroom_partner'], latin_partner=sql_dict['latin_partner'],
                            ballroom_role=sql_dict['ballroom_role'], latin_role=sql_dict['latin_role'])
                if city is None:
                    query += ' AND {team} != ?'.format(team=sql_dict['city'])
                else:
                    query += ' AND {team} = ?'.format(team=sql_dict['city'])
                    team = city
            # Try to find a partner for a beginner, beginner combination
            if all([ballroom_level == levels['beginners'], latin_level == levels['beginners'],
                    partner_id is None, number_of_potential_partners == 0]):
                potential_partners += cursor.execute(query, (levels['beginners'], '', ballroom_role, '', team))\
                    .fetchall()
                potential_partners += cursor.execute(query, ('', levels['beginners'], '', latin_role, team)).fetchall()
            # Try to find a partner for a levels['beginners'], Null combination
            if all([ballroom_level == levels['beginners'], latin_level == '',
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.\
                    execute(query, (levels['beginners'], levels['beginners'], ballroom_role, ballroom_role, team))\
                    .fetchall()
            # Try to find a partner for a Null, beginner combination
            if all([ballroom_level == '', latin_level == levels['beginners'],
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.\
                    execute(query, (levels['beginners'], levels['beginners'], latin_role, latin_role, team)).fetchall()
            # Try to find a partner for a breiten, breiten combination
            if all([ballroom_level == levels['breitensport'], latin_level == levels['breitensport'],
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.execute(query, (levels['breitensport'], '', ballroom_role, '', team))\
                    .fetchall()
                potential_partners += cursor.execute(query, ('', levels['breitensport'], '', latin_role, team))\
                    .fetchall()
                potential_partners += cursor.execute(query, (levels['breitensport'], levels['open_class'],
                                                             ballroom_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], levels['breitensport'],
                                                             ballroom_role, latin_role, team)).fetchall()
            # Try to find a partner for a breiten, Null combination
            if all([ballroom_level == levels['breitensport'], latin_level == '',
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.execute(query, (levels['breitensport'], levels['breitensport'],
                                                             ballroom_role, ballroom_role, team)).fetchall()
                potential_partners += cursor.execute(query, (levels['breitensport'], levels['open_class'],
                                                             ballroom_role, ballroom_role, team)).fetchall()
            # Try to find a partner for a Null, Breiten combination
            if all([ballroom_level == '', latin_level == levels['breitensport'],
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.execute(query, (levels['breitensport'], levels['breitensport'],
                                                             latin_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], levels['breitensport'],
                                                             latin_role, latin_role, team)).fetchall()
            # Try to find a partner for a Breiten, Open combination
            if all([ballroom_level == levels['breitensport'], latin_level == levels['open_class'],
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.execute(query, (levels['breitensport'], levels['breitensport'],
                                                             ballroom_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, (levels['breitensport'], '', ballroom_role, '', team))\
                    .fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], levels['open_class'],
                                                             ballroom_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, ('', levels['open_class'], '', latin_role, team)).fetchall()
            # Try to find a partner for a Open, Breiten combination
            if all([ballroom_level == levels['open_class'], latin_level == levels['breitensport'],
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.execute(query, (levels['breitensport'], levels['breitensport'],
                                                             ballroom_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, ('', levels['breitensport'], '', latin_role, team))\
                    .fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], levels['open_class'],
                                                             ballroom_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], '', ballroom_role, '', team))\
                    .fetchall()
            # Try to find a partner for a Open, Open combination
            if all([ballroom_level == levels['open_class'], latin_level == levels['open_class'],
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.execute(query, (levels['breitensport'], levels['open_class'],
                                                             ballroom_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], levels['breitensport'],
                                                             ballroom_role, latin_role, team)).fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], '', ballroom_role, '', team))\
                    .fetchall()
                potential_partners += cursor.execute(query, ('', levels['open_class'], '', latin_role, team)).fetchall()
            # Try to find a partner for a Open, Null combination
            if all([ballroom_level == levels['open_class'], latin_level == '',
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor.execute(query, (levels['open_class'], levels['breitensport'],
                                                             ballroom_role, ballroom_role, team)).fetchall()
                potential_partners += cursor\
                    .execute(query, (levels['open_class'], levels['open_class'], ballroom_role, ballroom_role, team))\
                    .fetchall()
            # Try to find a partner for a Null, Open combination
            if all([ballroom_level == '', latin_level == levels['open_class'],
                    number_of_potential_partners == 0, partner_id is None]):
                potential_partners += cursor\
                    .execute(query, (levels['breitensport'], levels['open_class'], latin_role, latin_role, team))\
                    .fetchall()
                potential_partners += cursor.execute(query, (levels['open_class'], levels['open_class'],
                                                             latin_role, latin_role, team)).fetchall()
        # If there is a potential partner, randomly select one
        if partner_id is None:
            number_of_potential_partners = len(potential_partners)
            if number_of_potential_partners > 0:
                random_num = randint(0, number_of_potential_partners - 1)
                partner_id = potential_partners[random_num][gen_dict['id']]
        if partner_id is None:
            status_print('Found no match for {id1}'.format(id1=identifier))
        elif partner_id is not None and ballroom_partner == '' and latin_partner == '':
            status_print('Matched {id1} and {id2} together'.format(id1=identifier, id2=partner_id))
        elif partner_id == identifier:
            status_print('TEMP')
        status_update()
    return partner_id


def create_pair(first_dancer, second_dancer, connection, cursor):
    """Creates a pair of the two selected dancers and writes their data away in partner lists"""
    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=signup_list, id=sql_dict['id'])
    if first_dancer != second_dancer:
        first_dancer = cursor.execute(query, (first_dancer,)).fetchone()
        second_dancer = cursor.execute(query, (second_dancer,)).fetchone()
        if first_dancer is None:
            first_dancer_ballroom_role = ''
            first_dancer_latin_role = ''
        else:
            first_dancer_ballroom_role = first_dancer[gen_dict['ballroom_role']]
            first_dancer_latin_role = first_dancer[gen_dict['latin_role']]
        if first_dancer_ballroom_role == roles['follow'] or first_dancer_latin_role == roles['follow']:
            first_dancer, second_dancer = second_dancer, first_dancer
        if first_dancer is None:
            first_dancer_id = ''
            first_dancer_team = ''
            first_dancer_ballroom_level = ''
            first_dancer_latin_level = ''
        else:
            first_dancer_id = first_dancer[gen_dict['id']]
            first_dancer_team = first_dancer[gen_dict['city']]
            first_dancer_ballroom_level = first_dancer[gen_dict['ballroom_level']]
            first_dancer_latin_level = first_dancer[gen_dict['latin_level']]
        if second_dancer is None:
            second_dancer_id = ''
            second_dancer_team = ''
            second_dancer_ballroom_level = ''
            second_dancer_latin_level = ''
        else:
            second_dancer_id = second_dancer[gen_dict['id']]
            second_dancer_team = second_dancer[gen_dict['city']]
            second_dancer_ballroom_level = second_dancer[gen_dict['ballroom_level']]
            second_dancer_latin_level = second_dancer[gen_dict['latin_level']]
        query = 'INSERT INTO {tn} ({lead}, {follow}, {lead_city}, {follow_city}, ' \
                '{ballroom_level_lead}, {ballroom_level_follow}, {latin_level_lead}, {latin_level_follow}) ' \
                'VALUES (?, ?, ?, ?, ?, ?, ?, ?)'\
            .format(tn=partners_list,
                    lead=partner_sql_dict['lead'], follow=partner_sql_dict['follow'],
                    lead_city=partner_sql_dict['city_lead'], follow_city=partner_sql_dict['city_follow'],
                    ballroom_level_lead=partner_sql_dict['ballroom_level_lead'],
                    ballroom_level_follow=partner_sql_dict['ballroom_level_follow'],
                    latin_level_lead=partner_sql_dict['latin_level_lead'],
                    latin_level_follow=partner_sql_dict['latin_level_follow'])
        cursor.execute(query, (first_dancer_id, second_dancer_id, first_dancer_team, second_dancer_team,
                               first_dancer_ballroom_level, second_dancer_ballroom_level,
                               first_dancer_latin_level, second_dancer_latin_level))
        query = 'INSERT INTO {tn} ({lead}, {follow}, {lead_city}, {follow_city}, ' \
                '{ballroom_level_lead}, {ballroom_level_follow}, {latin_level_lead}, {latin_level_follow}) ' \
                'VALUES (?, ?, ?, ?, ?, ?, ?, ?)' \
            .format(tn=ref_partner_list,
                    lead=partner_sql_dict['lead'], follow=partner_sql_dict['follow'],
                    lead_city=partner_sql_dict['city_lead'], follow_city=partner_sql_dict['city_follow'],
                    ballroom_level_lead=partner_sql_dict['ballroom_level_lead'],
                    ballroom_level_follow=partner_sql_dict['ballroom_level_follow'],
                    latin_level_lead=partner_sql_dict['latin_level_lead'],
                    latin_level_follow=partner_sql_dict['latin_level_follow'])
        cursor.execute(query, (first_dancer_id, second_dancer_id, first_dancer_team, second_dancer_team,
                               first_dancer_ballroom_level, second_dancer_ballroom_level,
                               first_dancer_latin_level, second_dancer_latin_level))
        query = 'DELETE FROM {tn} WHERE {id} = ?'.format(tn=selection_list, id=sql_dict['id'])
        cursor.executemany(query, [(first_dancer_id,), (second_dancer_id,)])
        connection.commit()


def move_selected_contestant(identifier, connection, cursor):
    """Moves dancer, given id, from the selection list to selected list"""
    if identifier is not None:
        query = 'SELECT * FROM {tn} WHERE {id} = ?'.format(tn=selected_list, id=sql_dict['id'])
        dancer_selected = cursor.execute(query, (identifier,)).fetchall()
        if len(dancer_selected) == 0:
            query = 'INSERT INTO {tn1} SELECT * FROM {tn2} WHERE id = ?;'.format(tn1=selected_list, tn2=signup_list)
            cursor.execute(query, (identifier,))
            query = 'DELETE FROM {tn} WHERE id = ?'.format(tn=selection_list)
            cursor.execute(query, (identifier,))
            connection.commit()
            status_print('Selected {} for the NTDS'.format(identifier))


def remove_selected_contestant(identifier, connection, cursor):
    """Temp"""
    if identifier is not None:
        query = 'SELECT * FROM {tn} WHERE {role} = ?'.format(tn=ref_partner_list, role=partner_sql_dict['lead'])
        couple = cursor.execute(query, (identifier,)).fetchall()
        role = ''
        if len(couple) == 0:
            query = 'SELECT * FROM {tn} WHERE {role} = ?'.format(tn=ref_partner_list, role=partner_sql_dict['follow'])
            couple = cursor.execute(query, (identifier,)).fetchall()
            if len(couple) != 0:
                role = roles['follow']
        elif len(couple) != 0:
            role = roles['lead']
        couple = couple[0]
        couple_id = couple[partner_dict['num']]
        if role == roles['lead']:
            query = 'UPDATE {tn} SET {role} = "", {city} = "", {ballroom_level} = "", ' \
                    '{latin_level} = "" WHERE {role} = ?' \
                .format(tn=ref_partner_list,
                        role=partner_sql_dict['lead'], city=partner_sql_dict['city_lead'],
                        ballroom_level=partner_sql_dict['ballroom_level_lead'],
                        latin_level=partner_sql_dict['latin_level_lead'])
            cursor.execute(query, (identifier,))
        elif role == roles['follow']:
            query = 'UPDATE {tn} SET {role} = "", {city} = "", {ballroom_level} = "", ' \
                    '{latin_level} = "" WHERE {role} = ?' \
                .format(tn=ref_partner_list,
                        role=partner_sql_dict['follow'], city=partner_sql_dict['city_follow'],
                        ballroom_level=partner_sql_dict['ballroom_level_follow'],
                        latin_level=partner_sql_dict['latin_level_follow'])
            cursor.execute(query, (identifier,))
        connection.commit()
        query = 'SELECT * FROM {tn} WHERE {num} = ?'.format(tn=ref_partner_list, num=partner_sql_dict['num'])
        couple = cursor.execute(query, (couple_id,)).fetchall()
        couple = couple[0]
        if couple[partner_dict['lead']] == '' and couple[partner_dict['follow']] == '':
            query = 'DELETE FROM {tn} WHERE {num} = ?'.format(tn=ref_partner_list, num=partner_sql_dict['num'])
            cursor.execute(query, (couple_id,))
            connection.commit()
        query = 'INSERT INTO {tn1} SELECT * FROM {tn2} WHERE {id} = ?;'\
            .format(tn1=selection_list, tn2=signup_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        query = 'DELETE FROM {tn} WHERE {id} = ?'.format(tn=selected_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        connection.commit()
        status_print('Removed {} from the NTDS selection.'.format(identifier))


def delete_selected_contestant(identifier, connection, cursor):
    """"Temp"""
    if identifier is not None:
        status_print('Contestant {num} cancelled his/her signup for the NTDS.'.format(num=identifier))
        # remove_selected_contestant(identifier, connection=connection, cursor=cursor)
        # query = 'DELETE FROM {tn} WHERE {id} = ?'.format(tn=selected_list, id=sql_dict['id'])
        # cursor.execute(query, (identifier,))
        query = 'DELETE FROM {tn} WHERE {id} = ?'.format(tn=selection_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        query = 'INSERT INTO {tn1} SELECT * FROM {tn2} WHERE {id} = ?;' \
            .format(tn1=cancelled_list, tn2=signup_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        connection.commit()


def reinstate_selected_contestant(identifier, connection, cursor):
    """"Temp"""
    if identifier is not None:
        status_print('Put contestant {num} back on the list to be eligible for selection.'.format(num=identifier))
        query = 'DELETE FROM {tn} WHERE {id} = ?'.format(tn=cancelled_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        query = 'INSERT INTO {tn1} SELECT * FROM {tn2} WHERE {id} = ?;' \
            .format(tn1=selection_list, tn2=signup_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        connection.commit()


def include_selected_contestant(identifier, connection, cursor):
    """"Temp"""
    if identifier is not None:
        status_print('Put contestant {num} on the list to be eligible for selection.'.format(num=identifier))
        query = 'DELETE FROM {tn} WHERE {id} = ?'.format(tn=backup_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        query = 'INSERT INTO {tn1} SELECT * FROM {tn2} WHERE {id} = ?;' \
            .format(tn1=selection_list, tn2=signup_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        connection.commit()


def exclude_selected_contestant(identifier, connection, cursor):
    """"Temp"""
    if identifier is not None:
        status_print('Put contestant {num} on the backup list.'.format(num=identifier))
        query = 'DELETE FROM {tn} WHERE {id} = ?'.format(tn=selection_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        query = 'INSERT INTO {tn1} SELECT * FROM {tn2} WHERE {id} = ?;' \
            .format(tn1=backup_list, tn2=signup_list, id=sql_dict['id'])
        cursor.execute(query, (identifier,))
        connection.commit()


def select_bulk(limit, connection, cursor, no_partner=False):
    """"Temp"""
    if limit > boundaries['max_contestants']:
        limit = boundaries['min_fixed_lion_contestants']
    query = 'SELECT * FROM {tn}'.format(tn=selection_list)
    available_dancers = cursor.execute(query).fetchall()
    number_of_available_dancers = len(available_dancers)
    if number_of_available_dancers > 0:
        random_order = random.sample(range(0, number_of_available_dancers), number_of_available_dancers)
        for n in range(len(random_order)):
            dancer = available_dancers[random_order[n]]
            dancer_id = dancer[gen_dict['id']]
            if dancer_id is not None:
                query = ' SELECT * FROM {tn} WHERE {id} = ?'.format(tn=selection_list, id=sql_dict['id'])
                dancer_available = cursor.execute(query, (dancer_id,)).fetchone()
                if dancer_available is not None:
                    partner_id = find_partner(dancer_id, connection=connection, cursor=cursor)
                    if (partner_id is not None and no_partner is False) or no_partner is True:
                        create_pair(dancer_id, partner_id, connection=connection, cursor=cursor)
                        move_selected_contestant(dancer_id, connection=connection, cursor=cursor)
                        move_selected_contestant(partner_id, connection=connection, cursor=cursor)
            query = 'SELECT * FROM {tn}'.format(tn=selected_list)
            number_of_selected_dancers = len(cursor.execute(query).fetchall())
            if number_of_selected_dancers >= limit:
                break
    connection.commit()
    reset_selection_tables(connection=connection, cursor=cursor)


def create_city_beginners_list(cities_list, connection, cursor):
    """"Temp"""
    for city in cities_list:
        query = 'SELECT * FROM {tn1} WHERE ({ballroom_level} = ? OR {latin_level} = ?) AND {team} = ?' \
            .format(tn1=selection_list, ballroom_level=sql_dict['ballroom_level'],
                    latin_level=sql_dict['latin_level'], team=sql_dict['city'])
        max_city_beginners = len(cursor.execute(query, (levels['beginners'], levels['beginners'], city)).fetchall())
        if max_city_beginners > boundaries['min_guaranteed_beginners']:
            max_city_beginners = boundaries['min_guaranteed_beginners']
        query = 'INSERT INTO {tn} VALUES (?, ?, ?)'.format(tn=fixed_beginners_list)
        cursor.execute(query, (city, 0, max_city_beginners))
    connection.commit()


def update_city_beginners(cities_list, connection, cursor):
    """"Temp"""
    for city in cities_list:
        query = 'SELECT * FROM {tn} WHERE {city_lead} LIKE ?' \
            .format(tn=partners_list, city_lead=partner_sql_dict['city_lead'])
        number_of_city_beginners = len(cursor.execute(query, (city,)).fetchall())
        query = 'SELECT * FROM {tn} WHERE {city_follow} LIKE ?' \
            .format(tn=partners_list, city_follow=partner_sql_dict['city_follow'])
        number_of_city_beginners += len(cursor.execute(query, (city,)).fetchall())
        query = 'UPDATE {tn} SET {num} = ? WHERE {city} = ?' \
            .format(tn=fixed_beginners_list, num=city_sql_dict['number_of_contestants'], city=sql_dict['city'])
        cursor.execute(query, (number_of_city_beginners, city))
    connection.commit()


def get_lions_query():
    """"Temp"""
    query = 'SELECT * FROM {tn1} WHERE {team} = ?' \
        .format(tn1=selection_list, team=sql_dict['city'])
    sql_filter = []
    for level in lion_participants:
        sql_filter.append(' ( {ballroom_level} = "' + level + '"' + ' OR {latin_level} = "' + level + '" )')
    query_extension = ' AND (' + ' OR '.join(sql_filter) + ' )'
    query_extension = query_extension\
        .format(ballroom_level=sql_dict['ballroom_level'], latin_level=sql_dict['latin_level'])
    query += query_extension
    return query


def create_city_lions_list(cities_list, connection, cursor):
    """"Temp"""
    for city in cities_list:
        query = get_lions_query()
        max_city_lions = len(cursor.execute(query, (city,)).fetchall())
        if max_city_lions > boundaries['min_fixed_lion_contestants']:
            max_city_lions = boundaries['min_fixed_lion_contestants']
        query = 'INSERT INTO {tn} VALUES (?, ?, ?)'.format(tn=fixed_lions_list)
        cursor.execute(query, (city, 0, max_city_lions))
    connection.commit()


def update_city_lions(cities_list, connection, cursor):
    """"Temp"""
    for city in cities_list:
        query = 'SELECT * FROM {tn} WHERE {city_lead} = ? '\
            .format(tn=partners_list, city_lead=partner_sql_dict['city_lead'])
        sql_filter = []
        for lvl in lion_participants:
            sql_filter.append('{ballroom_level_lead} = "{lvl}"'
                              .format(ballroom_level_lead=partner_sql_dict['ballroom_level_lead'], lvl=lvl))
        sql_filter.append('{ballroom_level_lead} LIKE "%"'
                          .format(ballroom_level_lead=partner_sql_dict['ballroom_level_lead']))
        query += ' AND (' + ' OR '.join(sql_filter) + ')'
        sql_filter = []
        for lvl in lion_participants:
            sql_filter.append('{ballroom_level_follow} = "{lvl}"'
                              .format(ballroom_level_follow=partner_sql_dict['ballroom_level_follow'], lvl=lvl))
        sql_filter.append('{ballroom_level_follow} LIKE "%"'
                          .format(ballroom_level_follow=partner_sql_dict['ballroom_level_follow']))
        query += ' AND (' + ' OR '.join(sql_filter) + ' )'
        number_of_city_lions = len(cursor.execute(query, (city,)).fetchall())
        #
        query = 'SELECT * FROM {tn} WHERE {city_follow} = ? '\
            .format(tn=partners_list, city_follow=partner_sql_dict['city_follow'])
        sql_filter = []
        for lvl in lion_participants:
            sql_filter.append('{ballroom_level_lead} = "{lvl}"'
                              .format(ballroom_level_lead=partner_sql_dict['ballroom_level_lead'], lvl=lvl))
        sql_filter.append('{ballroom_level_lead} LIKE "%"'
                          .format(ballroom_level_lead=partner_sql_dict['ballroom_level_lead']))
        query += ' AND (' + ' OR '.join(sql_filter) + ')'
        sql_filter = []
        for lvl in lion_participants:
            sql_filter.append('{ballroom_level_follow} = "{lvl}"'
                              .format(ballroom_level_follow=partner_sql_dict['ballroom_level_follow'], lvl=lvl))
        sql_filter.append('{ballroom_level_follow} LIKE "%"'
                          .format(ballroom_level_follow=partner_sql_dict['ballroom_level_follow']))
        query += ' AND (' + ' OR '.join(sql_filter) + ' )'
        number_of_city_lions += len(cursor.execute(query, (city,)).fetchall())
        query = 'UPDATE {tn} SET {num} = ? WHERE {city} = ?' \
            .format(tn=fixed_lions_list, num=city_sql_dict['number_of_contestants'], city=sql_dict['city'])
        cursor.execute(query, (number_of_city_lions, city))
    connection.commit()


def max_rc(direction, worksheet):
    """Finds the maximum number of rows or columns of a worksheet"""
    max_dir = 0
    if direction == 'row':
        while True:
            max_dir += 1
            if worksheet.cell(row=max_dir, column=1).value is None:
                max_dir -= 1
                break
    elif direction == 'col':
        while True:
            max_dir += 1
            if worksheet.cell(row=1, column=max_dir).value is None:
                max_dir -= 1
                break
    return max_dir


def create_tables(connection, cursor):
    """"Drops existing tables and creates new ones"""
    # Drop all existing tables (from previous run)
    tables_to_drop = [signup_list, selection_list, selected_list, cancelled_list, backup_list,
                      team_list, partners_list, ref_partner_list, fixed_beginners_list, fixed_lions_list]
    for table in tables_to_drop:
        query = drop_table_query.format(table)
        cursor.execute(query)
    # Create new tables
    dancer_list_tables = [signup_list, selection_list, selected_list, cancelled_list, backup_list]
    for table in dancer_list_tables:
        query = dancers_list_query.format(table)
        cursor.execute(query)
    query = team_list_query.format(team_list)
    cursor.execute(query)
    query = paren_table_query.format(partners_list)
    cursor.execute(query)
    query = paren_table_query.format(ref_partner_list)
    cursor.execute(query)
    query = city_list_query.format(fixed_beginners_list)
    cursor.execute(query)
    query = city_list_query.format(fixed_lions_list)
    cursor.execute(query)
    connection.commit()


def check_available_signup_sheets():
    """"Creates list of all competing cities"""
    # Empty list that will contain the signup sheet file names for each of the teams
    competing_cities_array = []
    number_of_signup_sheets_found = 0
    for k, v in participating_teams_dict.items():
        competing_cities_array.append([v['team_name'], v['city'], v['signup_sheet']])
    for row in competing_cities_array:
        if os.path.isfile(path=row[team_dict['signup_list']]):
            number_of_signup_sheets_found += 1
        else:
            status_print('')
            status_print('Did not find the signup sheet of team {team} in the same directory as the program.'
                         .format(team=row[team_dict['city']]))
    if len(competing_cities_array) == 0:
        status_print('')
        status_print('No teams have been found in the config files.')
        return False
    elif number_of_signup_sheets_found == len(competing_cities_array):
        return True
    else:
        status_print('')
        status_print('The signup sheet of one or more teams that are supposed to enter the competition could not be '
                     'found. Please place all of the signup sheets in the same folder as this program and try again.')
        return False


def create_competing_teams(connection, cursor):
    """"Creates list of all competing cities"""
    # Empty list that will contain the signup sheet file names for each of the teams
    competing_cities_array = []
    for k, v in participating_teams_dict.items():
        competing_cities_array.append([v['team_name'], v['city'], v['signup_sheet']])
    # Write signup sheets into database
    query = 'INSERT INTO {tn} VALUES (?, ?, ?)'.format(tn=team_list)
    for row in competing_cities_array:
        if os.path.isfile(path=row[team_dict['signup_list']]):
            cursor.execute(query, row)
    connection.commit()
    query = 'SELECT * FROM {tn} ORDER BY {team}'.format(tn=team_list, team=list(team_dict.keys())[0])
    competing_cities_array = cursor.execute(query).fetchall()
    return competing_cities_array


def get_competing_teams(cursor):
    """"Temp"""
    query = 'SELECT * FROM {tn} ORDER BY {team}'.format(tn=team_list, team=list(team_dict.keys())[0])
    return cursor.execute(query).fetchall()


def get_competing_cities(cursor):
    """"Temp"""
    competing_teams = get_competing_teams(cursor)
    competing_cities = [row[team_dict['city']] for row in competing_teams]
    return competing_cities


def reset_selection_tables(connection, cursor):
    """"Resets the partners_list table"""
    query = drop_table_query.format(partners_list)
    cursor.execute(query)
    connection.commit()
    query = paren_table_query.format(partners_list)
    cursor.execute(query)
    connection.commit()


def collect_city_overview(source_table, target_table, users, cursor, connection, collect_data=False):
    """"Prints an overview"""
    if source_table == selected_list:
        query = 'SELECT {city}, COUNT() FROM {tn} GROUP BY {city}'.format(tn=source_table, city=sql_dict['city'])
    else:
        query = 'SELECT * FROM {tn} ORDER BY {city}'.format(tn=source_table, city=sql_dict['city'])
    ordered_cities = cursor.execute(query).fetchall()
    # if gather_stats and collect_data:
    if collect_data:
        query = 'CREATE TABLE IF NOT EXISTS {tn} (id INTEGER PRIMARY KEY AUTOINCREMENT'.format(tn=target_table)
        for city in ordered_cities:
            query += ', {city} INT'.format(city=city[0])
        query += ')'
        cursor.execute(query)
        query = 'INSERT INTO {tn} ('.format(tn=target_table)
        for city in ordered_cities:
            query += '{city}, '.format(city=city[0])
        query = query[:-2] + ') VALUES (' + '?,'*len(ordered_cities)
        query = query[:-1] + ')'
        cursor.execute(query, tuple([city[1] for city in ordered_cities]))
    for city in ordered_cities:
        overview = 'Number of selected {users} from {city} is: {number}'\
            .format(city=city[0], number=city[1], users=users)
        status_print(overview)
    status_print('')
    connection.commit()


def export_excel_lists(cursor, timestamp, city=None):
    """"Creates an Excel file, containing a list of who got selected and who is on the waiting list.
    One file containing all the dancers is generated for the for the organization, as well as a separate for
    each cities' team captain, containing lists of only that city"""
    status_print('Exporting all selected dancers')
    if city is None:
        output_file = output_key['path'] + '_' + timestamp + '_Overview' + xlsx_ext
    else:
        output_file = output_key['path'] + '_' + timestamp + '_' + city + xlsx_ext
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = 'Selected'
    output_data = list()
    output_data.append(template_values)
    if city is None:
        query = 'SELECT * FROM {tn} ORDER BY {id}'.format(tn=selected_list, id=sql_dict['id'])
        selected_dancers = cursor.execute(query).fetchall()
    else:
        query = 'SELECT * FROM {tn} WHERE {city} = ? ORDER BY {id}'\
            .format(tn=selected_list, city=sql_dict['city'], id=sql_dict['id'])
        selected_dancers = cursor.execute(query, (city,)).fetchall()
    output_data.extend(selected_dancers)
    for row in range(len(output_data)):
        for column in range(len(template_values)):
            cell = worksheet.cell(row=row + 1, column=column + 1)
            cell.value = output_data[row][column]
    for column in range(len(template_values)):
        worksheet.column_dimensions[get_column_letter(column+1)].width = column_widths[column] + 0.71
    worksheet.row_dimensions[1].height = 75
    for column in range(len(template_values)):
        cell = worksheet.cell(row=1, column=column + 1)
        cell.alignment = Alignment(wrap_text=True)
    worksheet.freeze_panes = 'A2'
    workbook.save(output_file)
    status_print('Exporting all dancers on the waiting list')
    worksheet = workbook.create_sheet('Waiting list')
    output_data = list()
    output_data.append(template_values)
    if city is None:
        query = 'SELECT * FROM {tn} ORDER BY {id}'.format(tn=selection_list, id=sql_dict['id'])
        waiting_dancers = cursor.execute(query).fetchall()
    else:
        query = 'SELECT * FROM {tn} WHERE {city} = ? ORDER BY {id}'\
            .format(tn=selection_list, city=sql_dict['city'], id=sql_dict['id'])
        waiting_dancers = cursor.execute(query, (city,)).fetchall()
    output_data.extend(waiting_dancers)
    for row in range(len(output_data)):
        for column in range(len(template_values)):
            cell = worksheet.cell(row=row + 1, column=column + 1)
            cell.value = output_data[row][column]
    status_print('Saving output: "{file}"'.format(file=output_file))
    status_print('')
    for column in range(len(template_values)):
        worksheet.column_dimensions[get_column_letter(column+1)].width = column_widths[column] + 0.71
    worksheet.row_dimensions[1].height = 75
    for column in range(len(template_values)):
        cell = worksheet.cell(row=1, column=column + 1)
        cell.alignment = Alignment(wrap_text=True)
    worksheet.freeze_panes = 'A2'
    workbook.save(output_file)


def create_stats_file(cursor, timestamp):
    """"Temp"""
    status_print('Exporting statistics file.')
    stats_title = get_competing_cities(cursor)
    stats_title.insert(0, 'run #')
    output_file = statistics_key['path'] + '_' + timestamp + xlsx_ext
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = levels['beginners']
    query = 'SELECT * FROM {tn} ORDER by "ïd"'.format(tn=beginners_list)
    output_data = list()
    output_data.append(stats_title)
    output_data.extend(cursor.execute(query).fetchall())
    status_print('Exporting {}.'.format(levels['beginners']))
    for row in range(len(output_data)):
        for column in range(len(output_data[0])):
            cell = worksheet.cell(row=row + 1, column=column + 1)
            cell.value = output_data[row][column]
    workbook.save(output_file)
    worksheet = workbook.create_sheet(lions)
    query = 'SELECT * FROM {tn} ORDER by "ïd"'.format(tn=lions_list)
    output_data = list()
    output_data.append(stats_title)
    output_data.extend(cursor.execute(query).fetchall())
    status_print('Exporting {}.'.format(lions))
    for row in range(len(output_data)):
        for column in range(len(output_data[0])):
            cell = worksheet.cell(row=row + 1, column=column + 1)
            cell.value = output_data[row][column]
    workbook.save(output_file)
    worksheet = workbook.create_sheet(contestants)
    query = 'SELECT * FROM {tn} ORDER by "ïd"'.format(tn=contestants_list)
    output_data = list()
    output_data.append(stats_title)
    output_data.extend(cursor.execute(query).fetchall())
    status_print('Exporting {}.'.format(contestants))
    for row in range(len(output_data)):
        for column in range(len(output_data[0])):
            cell = worksheet.cell(row=row + 1, column=column + 1)
            cell.value = output_data[row][column]
    workbook.save(output_file)
    worksheet = workbook.create_sheet(individuals)
    query = 'SELECT max({id}) FROM {tn}'.format(tn=signup_list, id=sql_dict['id'])
    signed_contestants = cursor.execute(query).fetchone()[0]
    query = 'SELECT * FROM {tn} ORDER by "ïd"'.format(tn=individual_list)
    output_title = list()
    output_title.extend(list(range(0, signed_contestants + 1)))
    output_title = list(map(str, output_title))
    output_title = ['#'+x for x in output_title]
    output_title[0] = 'run #'
    output_data = list()
    output_data.append(output_title)
    output_data.extend(cursor.execute(query).fetchall())
    status_print('Exporting {}.'.format(individuals))
    for row in range(len(output_data)):
        for column in range(len(output_data[0])):
            cell = worksheet.cell(row=row + 1, column=column + 1)
            cell.value = output_data[row][column]
    workbook.save(output_file)
    output_db = statistics_key['path'] + '_' + timestamp + db_ext
    status_print('Exporting database.')
    status_print('')
    shutil.copy2(database_key['path'], output_db)
    status_print('Finished exporting statistics')
    # TODO add plots and statistics analysis


def print_ntds_config():
    """"Displays the boundary conditions of the tournament in the welcome text"""
    # if user_boundaries:
    #     status_print('Using user setting for boundary conditions.')
    # else:
    #     status_print('Using default settings.')
    status_print('')
    status_print('Contestant numbers for this NTDS selection:')
    status_print('')
    status_print('Maximum number of contestants: {num}'.format(num=boundaries['max_contestants']))
    status_print('Guaranteed beginners per team: {num}'.format(num=boundaries['min_guaranteed_beginners']))
    status_print('Guaranteed Lion contestants per team: {num}'.format(num=boundaries['min_fixed_lion_contestants']))
    status_print('Cutoff for selecting all beginners: {num}'.format(num=boundaries['beginner_signup_cutoff']))
    status_print('Buffer for selecting contestants at the end: {num}'.format(num=boundaries['buffer_for_selection']))
    msg = 'Levels participating for the Lion: '
    for level in lion_participants:
        msg += '{lvl}, '.format(lvl=level)
    msg = msg[:-2]
    status_print(msg)
    status_print('')


def welcome_text():
    """"Text displayed when opening the program for the first time."""
    status_text.config(state=NORMAL)
    status_text.config(wrap=WORD)
    status_text.config(state=DISABLED)
    status_print('Welcome to the NTDS Selection!')
    status_print('')
    status_print('You can start a new selection, update an existing selection, '
                 'or change the settings with the buttons in the bottom right corner.')
    status_print('')
    status_print('For an overview of the available commands, type "help" in the command prompt.')
    status_print('')
    status_text.config(state=NORMAL)
    status_text.config(wrap=NONE)
    status_text.config(state=DISABLED)
    data_text.config(state=NORMAL)
    data_text.delete('1.0', END)
    data_text.insert(END, 'Data about the number of contestants, First Aid Officers, required sleeping locations, etc. '
                          'will be displayed here once a database has been created/selected.')
    data_text.config(state=DISABLED)


def command_help_text():
    """"Help text"""
    status_print('')
    status_print('Listing all commands:')
    status_print('')
    status_print(list_selected)
    status_print('Lists all dancers that were selected for the tournament.')
    status_print('')
    status_print(list_selected + '_beginners / _breiten / _closed / _open')
    status_print('Lists all dancers of the chosen level that were selected for the tournament.')
    status_print('')
    status_print(list_available)
    status_print('Lists all dancers available for selection.')
    status_print('')
    status_print(list_available + '_beginners / _breiten / _closed / _open')
    status_print('Lists all dancers of the chosen level that are available for selection for the tournament.')
    status_print('')
    status_print(list_backup)
    status_print('Lists all dancers that are on the backup list for the tournament')
    status_print('')
    status_print(list_backup + '_beginners / _breiten / _closed / _open')
    status_print('Lists all dancers of the chosen level that are on the backup list for the tournament.')
    status_print('')
    status_print(list_cancelled)
    status_print('Lists all dancers that have cancelled their signup for the tournament.')
    status_print('')
    status_print(list_cv)
    status_print('Lists all dancers available for selection that want to volunteer at the tournament.')
    status_print('')
    status_print(list_pv)
    status_print('Lists all dancers available for selection that were a volunteer at a previous tournament.')
    status_print('')
    status_print(list_fa)
    status_print('Lists all dancers available for selection that are a qualified First Aid Officer.')
    status_print('')
    status_print(list_ero)
    status_print('Lists all dancers available for selection that are a qualified Emergency Response Officer.')
    status_print('')
    status_print(list_ballroom_jury)
    status_print('Lists all dancers available for selection that can be a Ballroom Jury.')
    status_print('')
    status_print(list_latin_jury)
    status_print('Lists all dancers available for selection that can be a Latin Jury.')
    status_print('')
    status_print(select + ' n')
    status_print('Selects contestant number "n" (and their signed partner) for the NTDS. '
                 '"n" can be a list of numbers to select a range of contestants at once.')
    status_print('')
    status_print(selectp + ' n')
    status_print('Selects contestant number "n", and a (virtual) partner for the NTDS.')
    status_print('')
    status_print(remove + ' n')
    status_print('Removes contestant number "n" from the selected contestants.')
    status_print('')
    status_print(removep + ' n')
    status_print('Removes contestant number "n", and their (virtual) partner from the selected contestants.')
    status_print('')
    status_print(delete + ' n')
    status_print('Cancels the signup of contestant number "n", removing them from the tournament selection pool.')
    status_print('')
    status_print(reinstate + ' n')
    status_print('Reinstates the signup of contestant number "n", placing them into the tournament selection pool.')
    status_print('')
    status_print(selectr + ' input')
    status_print('Selects a random contestant and a (virtual) partner for the NTDS.')
    status_print('"input" can be one of the following statements:')
    status_print('A city name. Doing this will select a random contestant from the given city.')
    status_print(' / '.join((rinput_fa, rinput_ero, rinput_ballroom_jury, rinput_latin_jury)))
    status_print('Select a random dancers that can volunteer as a First Aid Officer, an Emergency Response Officer, '
                 'a Ballroom jury or a Latin jury respectively.')
    status_print('')
    status_print(switchlf + ' n')
    status_print('Switches the roles from contestant number "n" from {lead}/{lead} to {follow}/{follow} or the other '
                 'way around. "n" can be a list of numbers to select a range of contestants at once.')
    status_print('')
    status_print(print_contestants)
    status_print('Prints a list of how much contestants each city has had selected.')
    status_print('')
    status_print(print_breakdown)
    status_print('Prints a list of how much contestants each city has had selected, '
                 'breaking it down per class and discipline.')
    status_print('')
    status_print(finish_selection)
    status_print('Finishes the NTDS selection by adding random contestants.')
    status_print('')
    status_print(export)
    status_print('Creates Excel files containing the selection data.')
    status_print('')
    status_print(import_backup)
    status_print('Imports a second wave of contestants into the database. These contestants will be placed in a '
                 'separate table.')
    status_print('')
    status_print(stats + ' n')
    status_print('Simulates the automatic selection process "n" times (default {runs}), '
                 'to gather statistical data about the selection process.'.format(runs=runs))
    status_print('Used to verify that the selection process is fair (random).')
    status_print('WARNING: Doing this might take up to 4 hours, longer if the machine running the simulations is being '
                 'actively used. ')
    status_print('')


def status_print(message, wrap=True):
    """"Prints the message passed to the program screen"""
    status_text.config(state=NORMAL)
    if wrap is True:
        message = textwrap.fill(message, status_text_width)
    status_text.insert(END, message)
    status_text.insert(END, '\n')
    status_text.update()
    status_text.see(END)
    status_text.config(state=DISABLED)


def print_table(table):
    """"Formatting for printing tables"""
    formatted_table = []
    for dancer in table:
        if dancer[gen_dict['ln_prefix']] == '':
            name = dancer[gen_dict['first_name']] + ' ' + dancer[gen_dict['last_name']]
        else:
            name = dancer[gen_dict['first_name']] + ' ' + dancer[gen_dict['ln_prefix']] + ' ' + \
                   dancer[gen_dict['last_name']]
        formatted_dancer = [str(dancer[gen_dict['id']]), name,
                            dancer[gen_dict['ballroom_level']], dancer[gen_dict['latin_level']],
                            str(dancer[gen_dict['ballroom_partner']]), str(dancer[gen_dict['latin_partner']]),
                            dancer[gen_dict['ballroom_role']], dancer[gen_dict['latin_role']],
                            dancer[gen_dict['ballroom_mandatory_blind_date']],
                            dancer[gen_dict['latin_mandatory_blind_date']],
                            dancer[gen_dict['first_aid']], dancer[gen_dict['emergency_response_officer']],
                            dancer[gen_dict['ballroom_jury']], dancer[gen_dict['latin_jury']],
                            dancer[gen_dict['student']], dancer[gen_dict['sleeping_location']],
                            dancer[gen_dict['current_volunteer']], dancer[gen_dict['past_volunteer']],
                            dancer[gen_dict['city']]]
        formatted_table.append(formatted_dancer)
    status_text.config(wrap=NONE)
    print_table_header = ['Id', 'Name', 'Ballroom level', 'Latin level', 'Ballroom partner', 'Latin partner',
                          'Ballroom role', 'Latin role',
                          'Ballroom mandatory blind date', 'Latin mandatory blind date',
                          'First Aid', 'Emergency Response Officer', 'Ballroom jury', 'Latin jury',
                          'Student', 'Sleeping location', 'Volunteer', 'Past volunteer', 'Team']
    status_print(tabulate(formatted_table, headers=print_table_header, tablefmt='grid'), wrap=False)


def status_update():
    """"Data on the contestants of the active database"""
    status_connection = sql.connect(database_key['path'])
    status_cursor = status_connection.cursor()
    query = 'SELECT * FROM {tn}'.format(tn=selected_list)
    status_dict['number_of_contestants'] = len(status_cursor.execute(query).fetchall())

    query = 'SELECT * FROM {tn} WHERE {ballroom_level} = ? AND {ballroom_role} = ?'\
        .format(tn=selected_list, ballroom_level=sql_dict['ballroom_level'], ballroom_role=sql_dict['ballroom_role'])
    status_dict['number_of_beginner_ballroom_leads'] = \
        len(status_cursor.execute(query, (levels['beginners'], roles['lead'])).fetchall())
    status_dict['number_of_breiten_ballroom_leads'] = \
        len(status_cursor.execute(query, (levels['breitensport'], roles['lead'])).fetchall())
    status_dict['number_of_closed_ballroom_leads'] = \
        len(status_cursor.execute(query, (levels['closed'], roles['lead'])).fetchall())
    status_dict['number_of_open_ballroom_leads'] = \
        len(status_cursor.execute(query, (levels['open_class'], roles['lead'])).fetchall())
    status_dict['number_of_beginner_ballroom_follows'] = \
        len(status_cursor.execute(query, (levels['beginners'], roles['follow'])).fetchall())
    status_dict['number_of_breiten_ballroom_follows'] = \
        len(status_cursor.execute(query, (levels['breitensport'], roles['follow'])).fetchall())
    status_dict['number_of_closed_ballroom_follows'] = \
        len(status_cursor.execute(query, (levels['closed'], roles['follow'])).fetchall())
    status_dict['number_of_open_ballroom_follows'] = \
        len(status_cursor.execute(query, (levels['open_class'], roles['follow'])).fetchall())

    query = 'SELECT * FROM {tn} WHERE {latin_level} = ? AND {latin_role} = ?' \
        .format(tn=selected_list, latin_level=sql_dict['latin_level'], latin_role=sql_dict['latin_role'])
    status_dict['number_of_beginner_latin_leads'] = \
        len(status_cursor.execute(query, (levels['beginners'], roles['lead'])).fetchall())
    status_dict['number_of_breiten_latin_leads'] = \
        len(status_cursor.execute(query, (levels['breitensport'], roles['lead'])).fetchall())
    status_dict['number_of_closed_latin_leads'] = \
        len(status_cursor.execute(query, (levels['closed'], roles['lead'])).fetchall())
    status_dict['number_of_open_latin_leads'] = \
        len(status_cursor.execute(query, (levels['open_class'], roles['lead'])).fetchall())
    status_dict['number_of_beginner_latin_follows'] = \
        len(status_cursor.execute(query, (levels['beginners'], roles['follow'])).fetchall())
    status_dict['number_of_breiten_latin_follows'] = \
        len(status_cursor.execute(query, (levels['breitensport'], roles['follow'])).fetchall())
    status_dict['number_of_closed_latin_follows'] = \
        len(status_cursor.execute(query, (levels['closed'], roles['follow'])).fetchall())
    status_dict['number_of_open_latin_follows'] = \
        len(status_cursor.execute(query, (levels['open_class'], roles['follow'])).fetchall())

    query = 'SELECT * FROM {tn} WHERE {first_aid} = ?'.format(tn=selected_list, first_aid=sql_dict['first_aid'])
    status_dict['number_of_first_aid_yes'] = \
        len(status_cursor.execute(query, (NTDS_options['first_aid']['yes'],)).fetchall())
    status_dict['number_of_first_aid_maybe'] = \
        len(status_cursor.execute(query, (NTDS_options['first_aid']['maybe'],)).fetchall())

    query = 'SELECT * FROM {tn} WHERE {first_aid} = ?'\
        .format(tn=selected_list, first_aid=sql_dict['emergency_response_officer'])
    status_dict['number_of_emergency_response_officer_yes'] = \
        len(status_cursor.execute(query, (NTDS_options['emergency_response_officer']['yes'],)).fetchall())
    status_dict['number_of_emergency_response_officer_maybe'] = \
        len(status_cursor.execute(query, (NTDS_options['emergency_response_officer']['maybe'],)).fetchall())

    query = 'SELECT * FROM {tn} WHERE ballroom_level = ? AND ballroom_mandatory_blind_date = ?' \
        .format(tn=selected_list, ballroom_level=sql_dict['ballroom_level'],
                ballroom_mandatory_blind_date=sql_dict['ballroom_mandatory_blind_date'])
    status_dict['number_of_mandatory_breiten_ballroom_blind_daters'] = \
        len(status_cursor.execute(query, (levels['breitensport'], options_ymn['yes'])).fetchall())

    query = 'SELECT * FROM {tn} WHERE latin_level = ? AND latin_mandatory_blind_date = ?' \
        .format(tn=selected_list, latin_level=sql_dict['latin_level'],
                latin_mandatory_blind_date=sql_dict['latin_mandatory_blind_date'])
    status_dict['number_of_mandatory_breiten_latin_blind_daters'] = \
        len(status_cursor.execute(query, (levels['breitensport'], options_ymn['yes'])).fetchall())

    query = 'SELECT * FROM {tn} WHERE {ballroom_jury} = ?'\
        .format(tn=selected_list, ballroom_jury=sql_dict['ballroom_jury'])
    status_dict['number_of_ballroom_jury_yes'] = \
        len(status_cursor.execute(query, (NTDS_options['ballroom_jury']['yes'],)).fetchall())
    status_dict['number_of_ballroom_jury_maybe'] = \
        len(status_cursor.execute(query, (NTDS_options['ballroom_jury']['maybe'],)).fetchall())

    query = 'SELECT * FROM {tn} WHERE {latin_jury} = ?'.format(tn=selected_list, latin_jury=sql_dict['latin_jury'])
    status_dict['number_of_latin_jury_yes'] = \
        len(status_cursor.execute(query, (NTDS_options['latin_jury']['yes'],)).fetchall())
    status_dict['number_of_latin_jury_maybe'] = \
        len(status_cursor.execute(query, (NTDS_options['latin_jury']['maybe'],)).fetchall())

    query = 'SELECT * FROM {tn} WHERE {current_volunteer} = ?'\
        .format(tn=selected_list, current_volunteer=sql_dict['current_volunteer'])
    status_dict['number_of_current_volunteer_yes'] = \
        len(status_cursor.execute(query, (NTDS_options['current_volunteer']['yes'],)).fetchall())
    status_dict['number_of_current_volunteer_maybe'] = \
        len(status_cursor.execute(query, (NTDS_options['current_volunteer']['maybe'],)).fetchall())

    query = 'SELECT * FROM {tn} WHERE {past_volunteer} = ?'\
        .format(tn=selected_list, past_volunteer=sql_dict['past_volunteer'])
    status_dict['number_of_past_volunteer'] = \
        len(status_cursor.execute(query, (NTDS_options['past_volunteer']['yes'],)).fetchall())

    query = 'SELECT * FROM {tn} WHERE {sleeping_location} = ?' \
        .format(tn=selected_list, sleeping_location=sql_dict['sleeping_location'])
    status_dict['number_of_sleeping_spots'] = \
        len(status_cursor.execute(query, (NTDS_options['sleeping_location']['yes'],)).fetchall())

    query = 'SELECT * FROM {tn} WHERE {student} = ?' \
        .format(tn=selected_list, student=sql_dict['student'])
    status_dict['number_of_students'] = \
        len(status_cursor.execute(query, (NTDS_options['student']['yes'],)).fetchall())

    data_text.config(state=NORMAL)
    data_text.delete('1.0', END)
    data_text.insert(END, 'Selected database name: {name}\n'.format(name=database_key['db'] + db_ext))
    data_text.insert(END, '\n')
    data_text.insert(END, 'Total number of contestants: {num}\n'.format(num=status_dict['number_of_contestants']))
    data_text.insert(END, '\n')
    if levels['beginners'] in classes:
        data_text.insert(END, 'Beginners Ballroom Leads: {num}\n'
                         .format(num=status_dict['number_of_beginner_ballroom_leads']))
        data_text.insert(END, 'Beginners Ballroom Follows: {num}\n'
                         .format(num=status_dict['number_of_beginner_ballroom_follows']))
        data_text.insert(END, 'Beginners Latin Leads: {num}\n'
                         .format(num=status_dict['number_of_beginner_latin_leads']))
        data_text.insert(END, 'Beginners Latin Follows: {num}\n'
                         .format(num=status_dict['number_of_beginner_latin_follows']))
        data_text.insert(END, '\n')
    if levels['breitensport'] in classes:
        data_text.insert(END, 'Breitensport Ballroom Leads: {num}\n'
                         .format(num=status_dict['number_of_breiten_ballroom_leads']))
        data_text.insert(END, 'Breitensport Ballroom Follows: {num}\n'
                         .format(num=status_dict['number_of_breiten_ballroom_follows']))
        data_text.insert(END, 'Breitensport Latin Leads: {num}\n'
                         .format(num=status_dict['number_of_breiten_latin_leads']))
        data_text.insert(END, 'Breitensport Latin Follows: {num}\n'
                         .format(num=status_dict['number_of_breiten_latin_follows']))
        data_text.insert(END, '\n')
    if levels['closed'] in classes:
        data_text.insert(END, 'CloseD Ballroom Leads: {num}\n'
                         .format(num=status_dict['number_of_closed_ballroom_leads']))
        data_text.insert(END, 'CloseD Ballroom Follows: {num}\n'
                         .format(num=status_dict['number_of_closed_ballroom_follows']))
        data_text.insert(END, 'CloseD Latin Leads: {num}\n'
                         .format(num=status_dict['number_of_closed_latin_leads']))
        data_text.insert(END, 'CloseD Latin Follows: {num}\n'
                         .format(num=status_dict['number_of_closed_latin_follows']))
        data_text.insert(END, '\n')
    if levels['open_class'] in classes:
        data_text.insert(END, 'Open Class Ballroom Leads: {num}\n'
                         .format(num=status_dict['number_of_open_ballroom_leads']))
        data_text.insert(END, 'Open Class Ballroom Follows: {num}\n'
                         .format(num=status_dict['number_of_open_ballroom_follows']))
        data_text.insert(END, 'Open Class Latin Leads: {num}\n'
                         .format(num=status_dict['number_of_open_latin_leads']))
        data_text.insert(END, 'Open Class Latin Follows: {num}\n'
                         .format(num=status_dict['number_of_open_latin_follows']))
        data_text.insert(END, '\n')
    data_text.insert(END, 'First Aid (Yes): {yes}\n'.format(yes=status_dict['number_of_first_aid_yes']))
    data_text.insert(END, 'First Aid (Maybe): {maybe}\n'.format(maybe=status_dict['number_of_first_aid_maybe']))
    data_text.insert(END, 'Emergency Response Officer (Yes): {yes}\n'
                     .format(yes=status_dict['number_of_emergency_response_officer_yes']))
    data_text.insert(END, 'Emergency Response Officer (Maybe): {maybe}\n'
                     .format(maybe=status_dict['number_of_emergency_response_officer_maybe']))
    data_text.insert(END, '\n')
    data_text.insert(END, 'Breitensport mandatory Ballroom Blind Daters: {num}\n'
                     .format(num=status_dict['number_of_mandatory_breiten_ballroom_blind_daters']))
    data_text.insert(END, 'Breitensport mandatory Latin Blind Daters: {num}\n'
                     .format(num=status_dict['number_of_mandatory_breiten_latin_blind_daters']))
    data_text.insert(END, '\n')
    data_text.insert(END, 'Ballroom juries (Yes): {yes}\n'.format(yes=status_dict['number_of_ballroom_jury_yes']))
    data_text.insert(END, 'Ballroom juries (Maybe): {maybe}\n'
                     .format(maybe=status_dict['number_of_ballroom_jury_maybe']))
    data_text.insert(END, 'Latin juries (Yes): {yes}\n'.format(yes=status_dict['number_of_latin_jury_yes']))
    data_text.insert(END, 'Latin juries (Maybe): {maybe}\n'.format(maybe=status_dict['number_of_latin_jury_maybe']))
    data_text.insert(END, '\n')
    data_text.insert(END, 'Volunteers (Yes): {yes}\n'.format(yes=status_dict['number_of_current_volunteer_yes']))
    data_text.insert(END, 'Volunteers (Maybe): {maybe}\n'
                     .format(maybe=status_dict['number_of_current_volunteer_maybe']))
    data_text.insert(END, 'Past volunteers: {yes}\n'.format(yes=status_dict['number_of_past_volunteer']))
    data_text.insert(END, '\n')
    data_text.insert(END, 'Sleeping spots: {yes}'.format(yes=status_dict['number_of_sleeping_spots']))
    data_text.insert(END, '\n')
    data_text.insert(END, 'Students: {yes}'.format(yes=status_dict['number_of_students']))
    data_text.see(END)
    data_text.config(state=DISABLED)
    status_cursor.close()
    status_connection.close()


def select_database(entry=None):
    """"Temp"""
    if entry is None:
        ask_database = EntryBox('Please give the database name', (database_key, 'db'))
        root.wait_window(ask_database.top)
    else:
        database_key['db'] = entry
    if database_key['db'].endswith(db_ext):
        database_key['path'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), database_key['db'])
    else:
        database_key['path'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), database_key['db'] + db_ext)
    database_key['session_timestamp'] = round(time.time())
    if os.path.isfile(path=database_key['path']):
        status_text.config(state=NORMAL)
        status_text.delete('1.0', END)
        status_text.config(state=DISABLED)
        status_print('Selected existing database: {name}'.format(name=database_key['path']))
        status_print('')
        cli_text.focus_set()
        sel_conn = sql.connect(database_key['path'])
        sel_curs = sel_conn.cursor()
        status_update()
        sel_curs.close()
        sel_conn.close()
    elif os.path.isfile(path=database_key['path']) is False and database_key['name'] != '':
        status_print('The file "{name}" does not exist.'.format(name=database_key['path']))
        status_print('')


def options_menu():
    """"Temp"""
    status_print('Work in progress...')


def list_list(table_list, cursor):
    query = 'SELECT * FROM {tn} ORDER BY {id}'.format(tn=table_list, id=sql_dict['id'])
    selected_contestants = cursor.execute(query).fetchall()
    if len(selected_contestants) > 0:
        status_print('')
        if table_list == selection_list:
            status_print('All {num} contestants that are available for selection for the NTDS:'
                         .format(num=len(selected_contestants)))
        elif table_list == selected_list:
            status_print('All {num} contestants that have been selected for the NTDS:'
                         .format(num=len(selected_contestants)))
        elif table_list == cancelled_list:
            status_print('All {num} contestants that have cancelled their signup for the NTDS:'
                         .format(num=len(selected_contestants)))
        elif table_list == backup_list:
            status_print('All {num} contestants that that are on the backup list for the NTDS:'
                         .format(num=len(selected_contestants)))
        status_print('')
        print_table(selected_contestants)
        status_print('')
    else:
        if table_list == selection_list:
            status_print('There are no contestants available for selection for the NTDS.')
        elif table_list == selected_list:
            status_print('There are no contestants selected for the NTDS.')
        elif table_list == cancelled_list:
            status_print('There are no contestants that have cancelled their signup for the NTDS.')
        elif table_list == backup_list:
            status_print('There are no contestants on the backup list for the NTDS.')
        status_print('')


def list_level(level, cursor, table_list=selection_list):
    """"Temp"""
    if level in classes:
        query = 'SELECT * FROM {tn} WHERE {ballroom_level} = ? OR {latin_level} = ? ORDER BY {id}' \
            .format(tn=table_list, ballroom_level=sql_dict['ballroom_level'],
                    latin_level=sql_dict['latin_level'], id=sql_dict['id'])
        available_contestants = cursor.execute(query, (level, level)).fetchall()
        if len(available_contestants) > 0:
            status_print('')
            if table_list == selection_list:
                status_print('All {num} {lvl} dancers that are available for selection for the NTDS:'
                             .format(num=len(available_contestants), lvl=level))
            elif table_list == selected_list:
                status_print('All {num} {lvl} dancers that are selected for the NTDS:'
                             .format(num=len(available_contestants), lvl=level))
            elif table_list == backup_list:
                status_print('All {num} {lvl} dancers that are on the backup list for the NTDS:'
                             .format(num=len(available_contestants), lvl=level))
            status_print('')
            print_table(available_contestants)
            status_print('')
        else:
            if table_list == selection_list:
                status_print('There are no {lvl} dancers available for selection for the NTDS'.format(lvl=level))
            elif table_list == selected_list:
                status_print('There are no {lvl} dancers that have been selected for the NTDS'.format(lvl=level))
            elif table_list == backup_list:
                status_print('There are no {lvl} dancers on the backup list for the NTDS'.format(lvl=level))
            status_print('')
    else:
        status_print('The {lvl} class is not participating in this tournament.'.format(lvl=level))


def list_volunteer(volunteer_role, cursor):
    """"Temp"""
    if volunteer_role == sql_dict['past_volunteer']:
        messages = ['',
                    'Contestants that were a volunteer at a previous tournament:',
                    'There are no contestants available for selection that volunteered at a previous tournament.']
    else:
        messages = ['Contestants that MIGHT want to volunteer ',
                    'Contestants that want to volunteer ',
                    'There are no contestants available for selection that want to volunteer ']
        volunteer_dict = {sql_dict['first_aid']: 'as a First Aid Officer:',
                          sql_dict['emergency_response_officer']: 'as an Emergency Response Officer:',
                          sql_dict['ballroom_jury']: 'as a Ballroom Jury:', sql_dict['latin_jury']: 'as a Latin Jury:',
                          sql_dict['current_volunteer']: '.'}
        messages = [message + volunteer_dict[volunteer_role] for message in messages]
    query = 'SELECT * FROM {tn} WHERE {v_role} = ? ORDER BY {id}' \
        .format(tn=selection_list, v_role=volunteer_role, id=sql_dict['id'])
    available_contestants = cursor.execute(query, (options_ymn['maybe'],)).fetchall()
    if len(available_contestants) > 0:
        status_print('')
        status_print(messages[0])
        status_print('')
        print_table(available_contestants)
        status_print('')
    available_contestants = cursor.execute(query, (options_ymn['yes'],)).fetchall()
    if len(available_contestants) > 0:
        status_print('')
        status_print(messages[1])
        status_print('')
        print_table(available_contestants)
        status_print('')
    else:
        status_print(messages[2])
        status_print('')


def select_random_volunteer(start_message, sql_var, cursor):
    """"Temp"""
    status_print(start_message)
    query = 'SELECT * from {tn} WHERE {user_input} = ?'.format(tn=selection_list, user_input=sql_var)
    dancers = cursor.execute(query, (options_ymn['yes'],)).fetchall()
    if dancers is not None:
        dancer = random.choice(dancers)
        dancer_id = dancer[gen_dict['id']]
        cli_parser('', alternate_input=selectp + ' {id}'.format(id=dancer_id))
    else:
        status_print('There are nog contestants that definitely want to fulfill this role.')
        status_print('Looking for contestants that might want to fulfill it.')
        dancers = cursor.execute(query, (options_ymn['maybe'],)).fetchall()
        if dancers is not None:
            dancer = random.choice(dancers)
            dancer_id = dancer[gen_dict['id']]
            cli_parser('', alternate_input=selectp + ' {id}'.format(id=dancer_id))
        else:
            status_print('There are no volunteers for this role to select')


# TODO create import back list option
def add_backup_list(backup_signup_list, city, connection, cursor):
    if os.path.isfile(path=backup_signup_list):
        query = 'SELECT max({id}) FROM {tn}'.format(id=sql_dict['id'], tn=signup_list)
        max_contestant_number = cursor.execute(query).fetchone()[0]
        # Get maximum number of rows and extract signup list
        workbook = openpyxl.load_workbook(backup_signup_list, data_only=True)
        worksheet = workbook.worksheets[0]
        max_r = max_rc('row', worksheet)
        city_signup_list = list(worksheet.iter_rows(min_col=1, min_row=2, max_col=max_col, max_row=max_r))
        # Convert data to 2d list, replace None values with an empty string,
        # increase the id numbers so that there are no duplicates, and add the city to the contestant
        city_signup_list = [[cell.value for cell in row] for row in city_signup_list]
        city_signup_list = [['' if elem is None else elem for elem in row] for row in city_signup_list]
        city_signup_list = [[elem + max_contestant_number if isinstance(elem, int) else elem for elem in row]
                            for row in city_signup_list]
        for row in city_signup_list:
            row.append(city)
        # Copy the contestants to the SQL database
        query = 'INSERT INTO {} VALUES ('.format(signup_list) + ('?,' * (max_col + 1))[:-1] + ');'
        for row in city_signup_list:
            cursor.execute(query, row)
        query = 'INSERT INTO {} VALUES ('.format(backup_list) + ('?,' * (max_col + 1))[:-1] + ');'
        for row in city_signup_list:
            cursor.execute(query, row)
        connection.commit()
        return len(city_signup_list)
    else:
        status_print('The file "{name}" does not exist.'.format(name=backup_signup_list))
        status_print('')
        return 0


# CLI commands
# Open commands, always available
echo = 'echo'
help_com = 'help'
exit_com = 'exit'
stats = '-stats'
db = '-db'
# Database commands, only available when a database is selected
select = '-select'
selectp = '-selectp'
remove = '-remove'
removep = '-removep'
delete = '-delete'
reinstate = '-reinstate'
selectr = '-selectr'
switchlf = '-switchlf'
include = '-include'
exclude = '-exclude'
list_selected = 'list_selected'
list_available = 'list_available'
list_cancelled = 'list_cancelled'
list_backup = 'list_backup'
list_selected_beginners = 'list_selected_beginners'
list_selected_breiten = 'list_selected_breiten'
list_selected_closed = 'list_selected_closed'
list_selected_open = 'list_selected_open'
list_available_beginners = 'list_available_beginners'
list_available_breiten = 'list_available_breiten'
list_available_closed = 'list_available_closed'
list_available_open = 'list_available_open'
list_backup_beginners = 'list_backup_beginners'
list_backup_breiten = 'list_backup_breiten'
list_backup_closed = 'list_backup_closed'
list_backup_open = 'list_backup_open'
list_cv = 'list_cv'
list_pv = 'list_pv'
list_fa = 'list_fa'
list_ero = 'list_ero'
list_ballroom_jury = 'list_ballroom_jury'
list_latin_jury = 'list_latin_jury'
print_contestants = 'print_contestants'
print_breakdown = 'print_breakdown'
finish_selection = 'finish_selection'
export = 'export'
create_stats = 'create_stats'
import_backup = 'import_backup_lists'
gen_sql = '-gen_sql'

rinput_fa = 'first_aid'
rinput_ero = 'ERO'
rinput_ballroom_jury = 'ballroom_jury'
rinput_latin_jury = 'latin_jury'


def cli_parser(event, alternate_input=''):
    """"Command line interface parser. Reads the given input and acts accordingly."""
    user_input = ''
    if alternate_input == '':
        command = cli_text.get()
        cli_text.delete(0, END)
    else:
        command = alternate_input

    connection = sql.connect(database_key['path'])
    cli_curs = connection.cursor()
    open_commands = [echo, help_com, exit_com,
                     stats, db]
    db_commands = [select, selectp, remove, removep, delete, reinstate, selectr, switchlf, include, exclude,
                   list_selected, list_available, list_cancelled, list_backup,
                   list_selected_beginners, list_selected_breiten, list_selected_closed, list_selected_open,
                   list_available_beginners, list_available_breiten, list_available_closed, list_available_open,
                   list_backup_beginners, list_backup_breiten, list_backup_closed, list_backup_open,
                   list_cv, list_pv, list_fa, list_ero, list_ballroom_jury, list_latin_jury,
                   print_contestants, print_breakdown,
                   finish_selection, export,
                   import_backup,
                   gen_sql]
    if command.startswith('-'):
        command = command.split(' ', 1)
        if len(command) > 1:
            user_input = command[1]
        else:
            user_input = ''
        command = command[0]
    if command in db_commands and os.path.isfile(database_key['path']) is True:
        if command == select:
            try:
                user_input = user_input.split(',')
                if len(user_input) > 1:
                    for ind, user in enumerate(user_input):
                        cli_parser('', alternate_input=select + ' {id}'.format(id=user_input[ind]))
                else:
                    user_input = user_input[0]
                    selected_id = int(user_input)
                    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selection_list, id=sql_dict['id'])
                    dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                    if dancer is not None:
                        ballroom_partner = dancer[gen_dict['ballroom_partner']]
                        latin_partner = dancer[gen_dict['latin_partner']]
                        if ballroom_partner != '' or latin_partner != '':
                            status_print('Dancer {num} signed with a partner, selecting both'.format(num=selected_id))
                            partner_id = find_partner(selected_id, connection=connection, cursor=cli_curs,
                                                      signed_partner_only=True)
                            create_pair(selected_id, partner_id, connection=connection, cursor=cli_curs)
                            move_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                            move_selected_contestant(partner_id, connection=connection, cursor=cli_curs)
                        else:
                            status_print('Selecting dancer number {num} on his/her own'.format(num=selected_id))
                            create_pair(selected_id, '', connection=connection, cursor=cli_curs)
                            move_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    else:
                        status_print('Dancer {num} is not on the selection list, '
                                     'and can therefor not be selected for the tournament'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == selectp:
            try:
                selected_id = int(user_input)
                query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selection_list, id=sql_dict['id'])
                dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                if dancer is not None:
                    partner_id = find_partner(selected_id, connection=connection, cursor=cli_curs)
                    create_pair(selected_id, partner_id, connection=connection, cursor=cli_curs)
                    move_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    move_selected_contestant(partner_id, connection=connection, cursor=cli_curs)
                else:
                    status_print('Dancer {num} is not on the selection list, '
                                 'and can therefor not be selected for the tournament'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == remove:
            try:
                user_input = user_input.split(',')
                if len(user_input) > 1:
                    for ind, user in enumerate(user_input):
                        cli_parser('', alternate_input=remove + ' {id}'.format(id=user_input[ind]))
                else:
                    user_input = user_input[0]
                    selected_id = int(user_input)
                    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selected_list, id=sql_dict['id'])
                    dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                    if dancer is not None:
                        remove_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    else:
                        status_print('Dancer {num} is not on the selected list, '
                                     'and can therefor not be removed from the tournament.'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == removep:
            try:
                selected_id = int(user_input)
                query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selected_list, id=sql_dict['id'])
                dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                if dancer is not None:
                    partner_id = ''
                    query = 'SELECT * FROM {tn} WHERE {role} = ?'\
                        .format(tn=ref_partner_list, role=partner_sql_dict['lead'])
                    couple = cli_curs.execute(query, (selected_id,)).fetchall()
                    role = ''
                    if len(couple) == 0:
                        query = 'SELECT * FROM {tn} WHERE {role} = ?'\
                            .format(tn=ref_partner_list, role=partner_sql_dict['follow'])
                        couple = cli_curs.execute(query, (selected_id,)).fetchall()
                        if len(couple) != 0:
                            role = roles['follow']
                    elif len(couple) != 0:
                        role = roles['lead']
                    couple = couple[0]
                    if role == roles['lead']:
                        partner_id = couple[partner_dict['follow']]
                    elif role == roles['follow']:
                        partner_id = couple[partner_dict['lead']]
                    if partner_id != '':
                        status_print('Removing dancers number {sel} and {par} from the NTDS.'
                                     .format(sel=selected_id, par=partner_id))
                    remove_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    remove_selected_contestant(partner_id, connection=connection, cursor=cli_curs)
                else:
                    status_print('Dancer {num} is not on the selected list, '
                                 'and can therefor not be removed from the tournament.'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == delete:
            try:
                user_input = user_input.split(',')
                if len(user_input) > 1:
                    for ind, user in enumerate(user_input):
                        cli_parser('', alternate_input=delete + ' {id}'.format(id=user_input[ind]))
                else:
                    user_input = user_input[0]
                    selected_id = int(user_input)
                    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selected_list, id=sql_dict['id'])
                    dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                    if dancer is not None:
                        remove_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                        delete_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    else:
                        query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selection_list, id=sql_dict['id'])
                        dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                        if dancer is not None:
                            delete_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                        else:
                            status_print('Dancer {num} is not on the selection list, '
                                         'and can therefor not be removed from the tournament.'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == reinstate:
            try:
                user_input = user_input.split(',')
                if len(user_input) > 1:
                    for ind, user in enumerate(user_input):
                        cli_parser('', alternate_input=reinstate + ' {id}'.format(id=user_input[ind]))
                else:
                    user_input = user_input[0]
                    selected_id = int(user_input)
                    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=cancelled_list, id=sql_dict['id'])
                    dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                    if dancer is not None:
                        reinstate_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    else:
                        status_print('Dancer {num} is not on the cancelled list, '
                                     'and is already a part of the tournament selection process.'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == include:
            try:
                user_input = user_input.split(',')
                if len(user_input) > 1:
                    for ind, user in enumerate(user_input):
                        cli_parser('', alternate_input=include + ' {id}'.format(id=user_input[ind]))
                else:
                    user_input = user_input[0]
                    selected_id = int(user_input)
                    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=backup_list, id=sql_dict['id'])
                    dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                    if dancer is not None:
                        include_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    else:
                        status_print('Dancer {num} is not on the backup list, '
                                     'and can therefor not be moved to the selection list.'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == exclude:
            try:
                user_input = user_input.split(',')
                if len(user_input) > 1:
                    for ind, user in enumerate(user_input):
                        cli_parser('', alternate_input=include + ' {id}'.format(id=user_input[ind]))
                else:
                    user_input = user_input[0]
                    selected_id = int(user_input)
                    query = 'SELECT * from {tn} WHERE {id} =?'.format(tn=selection_list, id=sql_dict['id'])
                    dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                    if dancer is not None:
                        exclude_selected_contestant(selected_id, connection=connection, cursor=cli_curs)
                    else:
                        status_print('Dancer {num} is not on the selection list, '
                                     'and can therefor not be moved to the backup list.'.format(num=selected_id))
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == selectr:
            # Create competing cities list
            cities = get_competing_cities(cursor=cli_curs)
            if user_input in cities:
                status_print('Selecting random contestant from ' + user_input + '.')
                query = 'SELECT * from {tn} WHERE {user_input} = ?'\
                    .format(tn=selection_list, user_input=sql_dict['city'])
                dancers = cli_curs.execute(query, (user_input,)).fetchall()
                if dancers is not None:
                    dancer = random.choice(dancers)
                    dancer_id = dancer[gen_dict['id']]
                    cli_parser('', alternate_input=selectp+' {id}'.format(id=dancer_id))
                else:
                    status_print('There are no dancers from {city} to select'.format(city=user_input))
            elif user_input == rinput_fa:
                select_random_volunteer('Selecting random First Aid volunteer.', sql_dict['first_aid'], cursor=cli_curs)
            elif user_input == rinput_ero:
                select_random_volunteer('Selecting random Emergency Response Officer volunteer.',
                                        sql_dict['emergency_response_officer'], cursor=cli_curs)
            elif user_input == rinput_ballroom_jury:
                select_random_volunteer('Selecting random Ballroom jury volunteer.',
                                        sql_dict['ballroom_jury'], cursor=cli_curs)
            elif user_input == rinput_latin_jury:
                select_random_volunteer('Selecting random Latin jury volunteer.',
                                        sql_dict['latin_jury'], cursor=cli_curs)
            elif user_input == "":
                status_print('Selecting random contestant from all available contestants.')
                query = 'SELECT * from {tn}'.format(tn=selection_list)
                dancers = cli_curs.execute(query).fetchall()
                if dancers is not None:
                    dancer = random.choice(dancers)
                    dancer_id = dancer[gen_dict['id']]
                    cli_parser('', alternate_input=selectp + ' {id}'.format(id=dancer_id))
                else:
                    status_print('There are no dancers left to choose from.')
            else:
                status_print(user_input + ' is not a valid input.')
        elif command == switchlf:
            try:
                user_input = user_input.split(',')
                if len(user_input) > 1:
                    for ind, user in enumerate(user_input):
                        cli_parser('', alternate_input=switchlf + ' {id}'.format(id=user_input[ind]))
                else:
                    user_input = user_input[0]
                    selected_id = int(user_input)
                    query = 'SELECT * from {tn} WHERE {id} = ?'.format(tn=signup_list, id=sql_dict['id'])
                    dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                    dancer_role = None
                    dancer_ballroom_role = None
                    if dancer is not None:
                        dancer_ballroom_role = dancer[gen_dict['ballroom_role']]
                        dancer_latin_role = dancer[gen_dict['latin_role']]
                        if dancer_ballroom_role == dancer_latin_role:
                            if dancer_ballroom_role == roles['lead']:
                                dancer_role = roles['follow']
                            elif dancer_ballroom_role == roles['follow']:
                                dancer_role = roles['lead']
                    if dancer_role is not None:
                        query = 'UPDATE {tn} SET {ballroom_role} = ?, {latin_role} = ? WHERE {id} = ?'\
                            .format(tn=signup_list, ballroom_role=sql_dict['ballroom_role'],
                                    latin_role=sql_dict['latin_role'], id=sql_dict['id'])
                        cli_curs.execute(query, (dancer_role, dancer_role, selected_id))
                        query = 'SELECT * from {tn} WHERE {id} = ?'.format(tn=selection_list, id=sql_dict['id'])
                        dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                        if dancer is not None:
                            query = 'UPDATE {tn} SET {ballroom_role} = ?, {latin_role} = ? WHERE {id} = ?' \
                                .format(tn=selection_list, ballroom_role=sql_dict['ballroom_role'],
                                        latin_role=sql_dict['latin_role'], id=sql_dict['id'])
                            cli_curs.execute(query, (dancer_role, dancer_role, selected_id))
                        query = 'SELECT * from {tn} WHERE {id} = ?'.format(tn=selected_list, id=sql_dict['id'])
                        dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                        if dancer is not None:
                            query = 'UPDATE {tn} SET {ballroom_role} = ?, {latin_role} = ? WHERE {id} = ?' \
                                .format(tn=selected_list, ballroom_role=sql_dict['ballroom_role'],
                                        latin_role=sql_dict['latin_role'], id=sql_dict['id'])
                            cli_curs.execute(query, (dancer_role, dancer_role, selected_id))
                        query = 'SELECT * from {tn} WHERE {id} = ?'.format(tn=backup_list, id=sql_dict['id'])
                        dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                        if dancer is not None:
                            query = 'UPDATE {tn} SET {ballroom_role} = ?, {latin_role} = ? WHERE {id} = ?' \
                                .format(tn=backup_list, ballroom_role=sql_dict['ballroom_role'],
                                        latin_role=sql_dict['latin_role'], id=sql_dict['id'])
                            cli_curs.execute(query, (dancer_role, dancer_role, selected_id))
                        query = 'SELECT * from {tn} WHERE {id} = ?'.format(tn=cancelled_list, id=sql_dict['id'])
                        dancer = cli_curs.execute(query, (selected_id,)).fetchone()
                        if dancer is not None:
                            query = 'UPDATE {tn} SET {ballroom_role} = ?, {latin_role} = ? WHERE {id} = ?' \
                                .format(tn=cancelled_list, ballroom_role=sql_dict['ballroom_role'],
                                        latin_role=sql_dict['latin_role'], id=sql_dict['id'])
                            cli_curs.execute(query, (dancer_role, dancer_role, selected_id))
                        status_print('Changed dancer number {id} roles from {before}/{before} to {after}/{after}.'
                                     .format(id=str(selected_id), before=dancer_ballroom_role, after=dancer_role))
                        connection.commit()
            except ValueError:
                status_print('No/Incorrect user number given.')
        elif command == gen_sql:
            status_print('Generating file...')
            query = 'SELECT * FROM {tn} ORDER BY {id}'.format(tn=selected_list, id=sql_dict['id'])
            selected_contestants = cli_curs.execute(query).fetchall()
            base_query = "INSERT INTO `person` SET `id` = '{id}', `fname` = '{fname}', `name` = '{lname}', `team` = " \
                         "(SELECT `id` FROM `team` WHERE `name` = \"{team_name} ({city})\");"
            competing_teams = get_competing_teams(cli_curs)
            base_query_couple = "INSERT INTO `start` (`tourn`,`lead`,`follow`) VALUES ({tourn}, {lead}, {follow});"
            with open('./populateContestants.sql', 'w', encoding='utf-8') as f1:
                f1.write('use `BADdb`;' + os.linesep + os.linesep)
                for contestant in selected_contestants:
                    f1.write(base_query.format(id=contestant[gen_dict['id']], fname=contestant[gen_dict['first_name']]
                                               .replace('ë', 'e'), lname=contestant[gen_dict['last_name']]
                                               .replace('ë', 'e'), city=contestant[gen_dict['city']],
                                               team_name=[item[0] for item in competing_teams if item[1] ==
                                                          contestant[gen_dict['city']]][0]) + os.linesep)
                query = 'SELECT * FROM {tn} WHERE {role} = ? AND {partner} != "" ORDER BY {id}'\
                    .format(tn=selected_list, role=sql_dict['ballroom_role'], partner=sql_dict['ballroom_partner'],
                            id=sql_dict['id'])
                selected_contestants = cli_curs.execute(query,(roles['lead'],)).fetchall()
                f1.write(os.linesep + os.linesep)
                for contestant in selected_contestants:
                    if contestant[gen_dict['ballroom_level']] == levels['beginners']:
                        tourn = 2
                    elif contestant[gen_dict['ballroom_level']] == levels['breitensport']:
                        tourn = 1
                    else:
                        tourn = 14
                    f1.write(base_query_couple
                             .format(lead=contestant[gen_dict['id']], tourn=tourn,
                                     follow=str(contestant[gen_dict['ballroom_partner']])) + os.linesep)
                query = 'SELECT * FROM {tn} WHERE {role} = ? AND {partner} != "" ORDER BY {id}' \
                    .format(tn=selected_list, role=sql_dict['latin_role'], partner=sql_dict['latin_partner'],
                            id=sql_dict['id'])
                selected_contestants = cli_curs.execute(query, (roles['lead'],)).fetchall()
                f1.write(os.linesep + os.linesep)
                for contestant in selected_contestants:
                    if contestant[gen_dict['latin_level']] == levels['beginners']:
                        tourn = 9
                    elif contestant[gen_dict['latin_level']] == levels['breitensport']:
                        tourn = 8
                    else:
                        tourn = 7
                    f1.write(base_query_couple
                             .format(lead=contestant[gen_dict['id']], tourn = tourn,
                                     follow=str(contestant[gen_dict['latin_partner']])) + os.linesep)
                f1.write('SHOW WARNINGS;')
            status_print('Generated populateContestants.sql\n')
        # List contestants (from a certain class) from specific lists
        elif command == list_selected:
            list_list(selected_list, cursor=cli_curs)
        elif command == list_available:
            list_list(selection_list, cursor=cli_curs)
        elif command == list_cancelled:
            list_list(cancelled_list, cursor=cli_curs)
        elif command == list_backup:
            list_list(backup_list, cursor=cli_curs)
        elif command == list_selected_beginners:
            list_level(levels['beginners'], cursor=cli_curs, table_list=selected_list)
        elif command == list_selected_breiten:
            list_level(levels['breitensport'], cursor=cli_curs, table_list=selected_list)
        elif command == list_selected_closed:
            list_level(levels['closed'], cursor=cli_curs, table_list=selected_list)
        elif command == list_selected_open:
            list_level(levels['open_class'], cursor=cli_curs, table_list=selected_list)
        elif command == list_available_beginners:
            list_level(levels['beginners'], cursor=cli_curs)
        elif command == list_available_breiten:
            list_level(levels['breitensport'], cursor=cli_curs)
        elif command == list_available_closed:
            list_level(levels['closed'], cursor=cli_curs)
        elif command == list_available_open:
            list_level(levels['open_class'], cursor=cli_curs)
        elif command == list_backup_beginners:
            list_level(levels['beginners'], cursor=cli_curs, table_list=backup_list)
        elif command == list_backup_breiten:
            list_level(levels['breitensport'], cursor=cli_curs, table_list=backup_list)
        elif command == list_backup_closed:
            list_level(levels['closed'], cursor=cli_curs, table_list=backup_list)
        elif command == list_backup_open:
            list_level(levels['open_class'], cursor=cli_curs, table_list=backup_list)
        # List volunteers (with a specific role)
        elif command == list_cv:
            list_volunteer(sql_dict['current_volunteer'], cursor=cli_curs)
        elif command == list_pv:
            list_volunteer(sql_dict['past_volunteer'], cursor=cli_curs)
        elif command == list_fa:
            list_volunteer(sql_dict['first_aid'], cursor=cli_curs)
        elif command == list_ero:
            list_volunteer(sql_dict['emergency_response_officer'], cursor=cli_curs)
        elif command == list_ballroom_jury:
            list_volunteer(sql_dict['ballroom_jury'], cursor=cli_curs)
        elif command == list_latin_jury:
            list_volunteer(sql_dict['latin_jury'], cursor=cli_curs)

        # Finish the selection of the tournament automatically
        elif command == finish_selection:
            status_print('')
            status_print('Finishing the selection for the NTDS automatically')
            select_bulk(boundaries['max_contestants']-1, connection=connection, cursor=cli_curs)
            # TODO add smart system to fill in gaps
            query = 'SELECT * FROM {tn}'.format(tn=selected_list)
            number_of_selected_dancers = len(cli_curs.execute(query).fetchall())
            if number_of_selected_dancers < boundaries['max_contestants']:
                select_bulk(boundaries['max_contestants'], connection=connection, cursor=cli_curs, no_partner=True)
        # Print number of contestants for each city
        elif command == print_contestants:
            status_print('')
            collect_city_overview(source_table=selected_list, target_table=contestants_list, users=contestants,
                                  cursor=cli_curs, connection=connection)
        elif command == print_breakdown:
            print_dict = {sql_dict['ballroom_level']: 'Ballroom', sql_dict['latin_level']: 'Latin'}
            query = 'SELECT {city}, COUNT() FROM {tn} GROUP BY {city}'.format(tn=selected_list, city=sql_dict['city'])
            ordered_cities = cli_curs.execute(query, ()).fetchall()
            query_entries = list()
            for cl in classes:
                query_entries.append([sql_dict['ballroom_level'], cl])
                query_entries.append([sql_dict['latin_level'], cl])
            for city in ordered_cities:
                city = city[0]
                query = 'SELECT * FROM {tn} WHERE {city} = ?'.format(tn=selected_list, city=sql_dict['city'])
                number_of_selected_city_dancers = len(cli_curs.execute(query, (city,)).fetchall())
                status_print('')
                status_print('Breakdown of the number of selected dancers from {city}:'.format(city=city))
                status_print('Total dancers selected:\t\t {num}'.format(num=number_of_selected_city_dancers))
                for entries in query_entries:
                    query = 'SELECT * FROM {tn} WHERE {temp} = ? AND {city} = ?' \
                        .format(tn=selected_list, city=sql_dict['city'], temp={})
                    query = query.format(entries[0])
                    number_of_selected_city_dancers = len(cli_curs.execute(query, (entries[1], city)).fetchall())
                    status_print('{level}, {division}:\t\t {num}'

                                 .format(division=print_dict[entries[0]], level=entries[1],
                                         num=number_of_selected_city_dancers))
        elif command == export:
            # Unix timestamp at time of exporting
            save_time = str(round(time.time()))
            # Export database (make copy)
            output_db = output_key['path'] + '_' + save_time + db_ext
            status_print('')
            status_print('Exporting database')
            status_print('Saving output: "{file}"'.format(file=output_db))
            status_print('')
            shutil.copy2(database_key['path'], output_db)
            # Export overview file
            status_print('Exporting overview file')
            export_excel_lists(cursor=cli_curs, timestamp=save_time)
            query = 'SELECT {city} FROM {tn} ORDER BY {city}'.format(tn=team_list, city=sql_dict['city'])
            all_cities = cli_curs.execute(query).fetchall()
            for city in all_cities:
                city = city[0]
                status_print('Exporting {city} overview file'.format(city=city))
                export_excel_lists(cursor=cli_curs, timestamp=save_time, city=city)
            status_print('Export complete.')
            status_print('Exported files can be found in the folder:')
            status_print('"{folder}"'.format(folder=output_key['path'].replace(output_key['name'], '')))
        elif command == create_stats:
            create_stats_file(cursor=cli_curs, timestamp=str(round(time.time())))
        # -import_backup_list NTDS_Groningen_Backup.xlsx Groningen
        elif command == import_backup:
            # TODO
            for k, v in participating_teams_dict.items():
                backup_sheet = v['signup_sheet'].replace(xlsx_ext, backup_ext)
                if os.path.exists(backup_sheet):
                    number_of_added_backup_dancers = add_backup_list(backup_sheet, v['city'],
                                                                     connection=connection, cursor=cli_curs)
                    status_print('Added ' + str(number_of_added_backup_dancers) + ' contestants from team ' + v['city']
                                 + '.')
                    move_used_signupsheet(backup_sheet)
                else:
                    status_print('Team ' + v['city'] + ' has no backup dancers.')
    if command in open_commands:
        if command == echo:
            status_print(command)
        elif command == help_com:
            command_help_text()
        elif command == exit_com:
            status_print('')
            status_print('Closing down program.')
            status_print('')
            time.sleep(0.5)
            root.destroy()
        elif command == stats:
            try:
                iterations = int(user_input)
            except ValueError:
                iterations = runs
            duration = time.time()
            timestamp = str(round(time.time()))
            status_path = status_key['path'] + '_' + timestamp + ini_ext
            main_selection()
            if iterations > 1:
                duration = int(time.time() - duration)
                duration_warning = 'One selection run took about {time} seconds.\n'.format(time=duration)
                duration_warning += 'Running all of the selections will take approximately {time} minutes.\n'\
                    .format(time=round(iterations*duration/60))
                duration_warning += 'Proceed?'
                if not os.path.isfile(path=status_path):
                    open(status_path, 'w+').close()
                cp = configparser.ConfigParser()
                if os.path.isfile(path=status_path):
                    add_config(cp, '1', status_dict)
                    with open(status_path, 'w') as cf:
                        cp.write(cf)
                if messagebox.askyesno(root, message=duration_warning):
                    for i in range(iterations-1):
                        status_print('')
                        status_print('Starting run {i} of {it}'.format(i=i+2, it=iterations))
                        status_print('')
                        main_selection()
                        if os.path.isfile(path=status_path):
                            add_config(cp, str(i+2), status_dict)
                            with open(status_path, 'w') as cf:
                                cp.write(cf)
                    connection = sql.connect(database_key['path'])
                    cli_curs = connection.cursor()
                    create_stats_file(cursor=cli_curs, timestamp=timestamp)
                else:
                    status_print('Cancelling statistics gathering.')
            # gather_stats = False
        elif command == db:
            selected_db = str(user_input)
            select_database(entry=selected_db)
    if command != '' and command not in open_commands and command not in db_commands:
        status_print('Unknown command: "{command}"'.format(command=command))
    elif command in db_commands and os.path.isfile(database_key['path']) is False:
        status_print('Command not available, no open database')
    if os.path.isfile(database_key['path']) is True and root.winfo_exists() == 1:
        status_update()
    cli_curs.close()
    connection.close()
    return event


def main_selection():
    ####################################################################################################################
    # Extract data from sign-up sheets
    ####################################################################################################################
    # Disable the CLI
    cli_text.config(state=DISABLED)
    # Time at which the run started
    start_time = time.time()
    # Connect to database and create a cursor
    status_print('')
    # Exit loop if no signup sheets are available
    if check_available_signup_sheets() is False:
        cli_text.config(state=NORMAL)
        return None
    status_print('Creating new database...')
    status_print('Database name: {name}'.format(name=default_db_name+db_ext))
    database_key['db'] = default_db_name
    database_key['path'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), database_key['db'] + db_ext)
    database_key['session_timestamp'] = round(time.time())
    conn = sql.connect(database_key['path'])
    curs = conn.cursor()
    # Create SQL tables
    status_print('')
    status_print('Creating tables...')
    create_tables(connection=conn, cursor=curs)
    # Create competing cities list
    competing_teams = create_competing_teams(connection=conn, cursor=curs)
    competing_cities = get_competing_cities(cursor=curs)
    # Copy the signup list from every team into the SQL database
    total_signup_list = list()
    total_number_of_contestants = 0
    status_print('')
    status_print('Collecting signup data...')
    for team in competing_teams:
        city = team[team_dict['city']]
        team_signup_list = team[team_dict['signup_list']]
        if os.path.isfile(path=team_signup_list):
            status_print('{city} is entering the tournament, adding contestants to signup list'.format(city=city))
            # Get maximum number of rows and extract signup list
            workbook = openpyxl.load_workbook(team_signup_list, data_only=True)
            worksheet = workbook.worksheets[0]
            max_r = max_rc('row', worksheet)
            city_signup_list = list(worksheet.iter_rows(min_col=1, min_row=2, max_col=max_col, max_row=max_r))
            # Convert data to 2d list, replace None values with an empty string,
            # increase the id numbers so that there are no duplicates, and add the city to the contestant
            city_signup_list = [[cell.value for cell in row] for row in city_signup_list]
            city_signup_list = [['' if elem is None else elem for elem in row] for row in city_signup_list]
            city_signup_list = [[elem + total_number_of_contestants if isinstance(elem, int) else elem for elem in row]
                                for row in city_signup_list]
            for row in city_signup_list:
                row.append(city)
            total_signup_list.extend(city_signup_list)
            # Copy the contestants to the SQL database
            query = 'INSERT INTO {} VALUES ('.format(signup_list) + ('?,' * (max_col + 1))[:-1] + ');'
            for row in city_signup_list:
                curs.execute(query, row)
            query = 'INSERT INTO {} VALUES ('.format(selection_list) + ('?,' * (max_col + 1))[:-1] + ');'
            for row in city_signup_list:
                curs.execute(query, row)
            total_number_of_contestants += len(city_signup_list)
            # move_used_signupsheet(team_signup_list)
        else:
            status_print('{city} is not entering the tournament'.format(city=city))
        conn.commit()
    conn.commit()

    ####################################################################################################################
    # Select the team captains and (virtual) partners
    ####################################################################################################################
    status_print('')
    status_print('Selecting team captains...')
    query = 'WITH cities AS (SELECT DISTINCT {city} FROM {tn1} ORDER BY RANDOM()) SELECT {tn1}.* FROM cities ' \
            'LEFT JOIN {tn1} ON ({tn1}.{city} = cities.{city} AND {tn1}.{tc} = "Ja") ORDER BY {tn1}.{tc}'\
        .format(tn1=selection_list, tc=sql_dict['team_captain'], city=sql_dict['city'])
    team_captains = curs.execute(query).fetchall()
    for captain in team_captains:
        captain_id = captain[gen_dict['id']]
        query = 'SELECT * FROM {tn1} WHERE {id} = ?'.format(tn1=selected_list, id=sql_dict['id'])
        captain_selected = curs.execute(query, (captain_id,)).fetchone()
        if captain_selected is None:
            partner_id = find_partner(captain_id, connection=conn, cursor=curs)
            if partner_id is None:
                status_print('No partner found for contestant number {num}.'.format(num=captain_id))
                status_print('Selecting contestant without a partner because he/she is a team captain.')
            create_pair(captain_id, partner_id, connection=conn, cursor=curs)
            move_selected_contestant(captain_id, connection=conn, cursor=curs)
            move_selected_contestant(partner_id, connection=conn, cursor=curs)
    conn.commit()
    query = drop_table_query.format(partners_list)
    curs.execute(query)
    query = paren_table_query.format(partners_list)
    curs.execute(query)

    ####################################################################################################################
    # Select beginners if less people have signed than the given cutoff
    ####################################################################################################################
    query = 'SELECT * FROM {tn1} WHERE {ballroom_level} = ? OR {latin_level} = ?' \
        .format(tn1=selection_list, ballroom_level=sql_dict['ballroom_level'], latin_level=sql_dict['latin_level'])
    all_beginners = curs.execute(query, (levels['beginners'], levels['beginners'])).fetchall()
    number_of_signed_beginners = len(all_beginners)
    if number_of_signed_beginners <= boundaries['beginner_signup_cutoff']:
        status_print('')
        status_print('Less than {num} Beginners signed up.'.format(num=boundaries['beginner_signup_cutoff']+1))
        status_print('Matching up as much couples as possible and selecting everyone...')
        create_city_beginners_list(competing_cities, connection=conn, cursor=curs)
        for beg in all_beginners:
            beg_id = beg[gen_dict['id']]
            query = 'SELECT * FROM {tn1} WHERE {id} = ?'.format(tn1=selected_list, id=sql_dict['id'])
            beginner_selected = curs.execute(query, (beg_id,)).fetchone()
            if beginner_selected is None:
                partner_id = find_partner(beg_id, connection=conn, cursor=curs)
                create_pair(beg_id, partner_id, connection=conn, cursor=curs)
                move_selected_contestant(beg_id, connection=conn, cursor=curs)
                move_selected_contestant(partner_id, connection=conn, cursor=curs)
                update_city_beginners(competing_cities, connection=conn, cursor=curs)
        conn.commit()
        reset_selection_tables(connection=conn, cursor=curs)

    ####################################################################################################################
    # Select beginners if more people have signed than the given cutoff
    ####################################################################################################################
    if number_of_signed_beginners > boundaries['beginner_signup_cutoff']:
        status_print('')
        status_print('More than {num} Beginners signed up.'.format(num=boundaries['beginner_signup_cutoff']))
        status_print('Selecting guaranteed beginners for each team...')
        create_city_beginners_list(competing_cities, connection=conn, cursor=curs)
        query = 'SELECT sum({max_beg})-sum({num}) FROM {tn}'\
            .format(tn=fixed_beginners_list, max_beg=city_sql_dict['max_contestants'],
                    num=city_sql_dict['number_of_contestants'])
        max_iterations = curs.execute(query).fetchone()[0]
        for iteration in range(max_iterations):
            query = 'SELECT * FROM {tn} ORDER BY {num}, RANDOM()'\
                .format(tn=fixed_beginners_list, num=city_sql_dict['number_of_contestants'])
            ordered_cities = curs.execute(query).fetchall()
            selected_city = None
            for city in ordered_cities:
                if selected_city is None:
                    number_of_selected_city_beginners = city[city_dict['number_of_contestants']]
                    max_number_of_selected_city_beginners = city[city_dict['max_contestants']]
                    if (max_number_of_selected_city_beginners - number_of_selected_city_beginners) > 0:
                        selected_city = city[city_dict[sql_dict['city']]]
            if selected_city is not None:
                query = 'SELECT * FROM {tn1} WHERE ({ballroom_level} = ? OR {latin_level} = ?) AND {city} = ?' \
                    .format(tn1=selection_list, ballroom_level=sql_dict['ballroom_level'],
                            latin_level=sql_dict['latin_level'], city=sql_dict['city'])
                selected_city_beginners = curs.\
                    execute(query, (levels['beginners'], levels['beginners'], selected_city,)).fetchall()
                number_of_city_beginners = len(selected_city_beginners)
                if number_of_city_beginners > 0:
                    random_order = random.sample(range(0, number_of_city_beginners), number_of_city_beginners)
                    partner_id = None
                    for order_city in ordered_cities:
                        if partner_id is None:
                            order_city = order_city[city_dict[sql_dict['city']]]
                            if order_city != selected_city:
                                query = 'SELECT * FROM {tn1} WHERE ({ballroom_level} = ? OR {latin_level} = ?) ' \
                                        'AND {city} = ?'\
                                    .format(tn1=selection_list, ballroom_level=sql_dict['ballroom_level'],
                                            latin_level=sql_dict['latin_level'], city=sql_dict['city'])
                                order_city_beginners = curs.\
                                    execute(query, (levels['beginners'], levels['beginners'], order_city,)).fetchall()
                                number_of_available_beginners = len(order_city_beginners)
                                if number_of_available_beginners > 0:
                                    for n in random_order:
                                        if partner_id is None:
                                            beg = selected_city_beginners[n]
                                            beginner_id = beg[gen_dict['id']]
                                            query = ' SELECT * FROM {tn} WHERE {id} = ?'\
                                                .format(tn=selected_list, id=sql_dict['id'])
                                            beginner_available = curs.execute(query, (beginner_id,)).fetchone()
                                            if beginner_available is None:
                                                partner_id = find_partner(beginner_id, connection=conn, cursor=curs,
                                                                          city=order_city)
                                                if partner_id is not None:
                                                    create_pair(beginner_id, partner_id, connection=conn, cursor=curs)
                                                    move_selected_contestant(beginner_id, connection=conn, cursor=curs)
                                                    move_selected_contestant(partner_id, connection=conn, cursor=curs)
                                                    update_city_beginners(competing_cities, connection=conn,
                                                                          cursor=curs)
        # Select beginner that were guaranteed entry, but could not be matched due to lack of partners
        query = 'SELECT * FROM {tn} ORDER BY {num}, RANDOM()'\
            .format(tn=fixed_beginners_list, num=city_sql_dict['number_of_contestants'])
        ordered_cities = curs.execute(query).fetchall()
        for iteration in range(len(ordered_cities) * boundaries['min_guaranteed_beginners']):
            query = 'SELECT * FROM {tn} ORDER BY {num}, RANDOM()'\
                .format(tn=fixed_beginners_list, num=city_sql_dict['number_of_contestants'])
            ordered_cities = curs.execute(query).fetchall()
            for city in ordered_cities:
                number_of_city_beginners = city[city_dict['number_of_contestants']]
                max_number_of_city_beginners = city[city_dict['max_contestants']]
                city = city[city_dict[sql_dict['city']]]
                if (max_number_of_city_beginners - number_of_city_beginners) > 0:
                    query = 'SELECT * FROM {tn1} WHERE ({ballroom_level} = ? OR {latin_level} = ?) AND {city} = ?' \
                        .format(tn1=selection_list, ballroom_level=sql_dict['ballroom_level'],
                                latin_level=sql_dict['latin_level'], city=sql_dict['city'])
                    city_beginners = curs.execute(query, (levels['beginners'], levels['beginners'], city,)).fetchall()
                    number_of_city_beginners = len(city_beginners)
                    if number_of_city_beginners > 0:
                        random_order = random.sample(range(0, number_of_city_beginners), number_of_city_beginners)
                        beg = city_beginners[random_order[0]]
                        beginner_id = beg[gen_dict['id']]
                        partner_id = find_partner(beginner_id, connection=conn, cursor=curs)
                        create_pair(beginner_id, partner_id, connection=conn, cursor=curs)
                        move_selected_contestant(beginner_id, connection=conn, cursor=curs)
                        move_selected_contestant(partner_id, connection=conn, cursor=curs)
                        update_city_beginners(competing_cities, connection=conn, cursor=curs)
        reset_selection_tables(connection=conn, cursor=curs)

    ####################################################################################################################
    # Select guaranteed lions contestants
    ####################################################################################################################
    status_print('')
    status_print('Selecting guaranteed lions for each team...')
    create_city_lions_list(competing_cities, connection=conn, cursor=curs)
    query = 'SELECT sum({max_lion})-sum({num}) FROM {tn}' \
        .format(tn=fixed_lions_list, max_lion=city_sql_dict['max_contestants'],
                num=city_sql_dict['number_of_contestants'])
    max_iterations = curs.execute(query).fetchone()[0]
    block_control = 0 # CONTROL fix
    error_city = 0 # CONTROL fix
    for iteration in range(max_iterations):
        query = 'SELECT * FROM {tn} ORDER BY {num}, RANDOM()'.format(tn=fixed_lions_list,
                                                                     num=city_sql_dict['number_of_contestants'])
        ordered_cities = curs.execute(query).fetchall()
        selected_city = None
        for city in ordered_cities:
            if selected_city is None:
                number_of_selected_city_lions = city[city_dict['number_of_contestants']]
                max_number_of_city_lions = city[city_dict['max_contestants']]
                if block_control == len(ordered_cities): # CONTROL fix
                    if city[0] == error_city: # CONTROL fix
                        number_of_selected_city_lions = max_number_of_city_lions # CONTROL fix
                        block_control = 0
                else: # CONTROL fix
                    if(max_number_of_city_lions - number_of_selected_city_lions) > 0:
                        selected_city = city[city_dict[sql_dict['city']]]
        if selected_city is not None:
            query = get_lions_query()
            selected_city_lions = curs.execute(query, (selected_city,)).fetchall()
            number_of_city_lions = len(selected_city_lions)
            if number_of_city_lions > 0:
                random_order = random.sample(range(0, number_of_city_lions), number_of_city_lions)
                partner_id = None
                block_control = 0 # CONTROL fix
                for order_city in ordered_cities:
                    if partner_id is None:
                        # TODO dit verwijdern CONTROL FIX
                        block_control += 1 # CONTROL fix
                        error_city = selected_city # CONTROL fix
                        order_city = order_city[city_dict[sql_dict['city']]]
                        if order_city != selected_city:
                            query = get_lions_query()
                            order_city_lions = curs.execute(query, (order_city,)).fetchall()
                            number_of_available_lions = len(order_city_lions)
                            if number_of_available_lions > 0:
                                for n in random_order:
                                    if partner_id is None:
                                        lion = selected_city_lions[n]
                                        lion_id = lion[gen_dict['id']]
                                        query = ' SELECT * FROM {tn} WHERE {id} = ?' \
                                            .format(tn=selected_list, id=sql_dict['id'])
                                        lion_available = curs.execute(query, (lion_id,)).fetchone()
                                        if lion_available is None:
                                            partner_id = find_partner(lion_id, connection=conn, cursor=curs,
                                                                      city=order_city)
                                            if partner_id is not None:
                                                create_pair(lion_id, partner_id, connection=conn, cursor=curs)
                                                move_selected_contestant(lion_id, connection=conn, cursor=curs)
                                                move_selected_contestant(partner_id, connection=conn, cursor=curs)
                                                update_city_lions(competing_cities, connection=conn, cursor=curs)
    reset_selection_tables(connection=conn, cursor=curs)

    ####################################################################################################################
    # Select remaining contestants
    ####################################################################################################################
    status_print('')
    status_print('Selecting the bulk of contestants that have signed up...')
    select_bulk(limit=boundaries['max_contestants'] - boundaries['buffer_for_selection'], connection=conn, cursor=curs)
    status_print('')
    status_print("--- Done in %.3f seconds ---" % (time.time() - start_time))
    status_print('')

    ####################################################################################################################
    # Collect user data from main selection
    ####################################################################################################################
    collect_city_overview(source_table=fixed_beginners_list, target_table=beginners_list, users=levels['beginners'],
                          cursor=curs, connection=conn, collect_data=True)
    collect_city_overview(source_table=fixed_lions_list, target_table=lions_list, users=lions,
                          cursor=curs, connection=conn, collect_data=True)
    collect_city_overview(source_table=selected_list, target_table=contestants_list, users=contestants,
                          cursor=curs, connection=conn, collect_data=True)
    # if gather_stats:
    query = 'SELECT * FROM {tn}'.format(tn=signup_list)
    all_dancers = curs.execute(query).fetchall()
    query = 'SELECT name FROM sqlite_master WHERE type = ? AND name = ?'
    individual_table_exists = len(curs.execute(query, ('table', individual_list)).fetchall())
    if individual_table_exists == 0:
        query = 'CREATE TABLE IF NOT EXISTS {tn} ({run} INTEGER PRIMARY KEY, '\
            .format(tn=individual_list, run=sql_run)
        for dancer in all_dancers:
            dancer_id = dancer[gen_dict['id']]
            query += ('"' + str(dancer_id) + '" INT, ')
        query = query[:-2]
        query += ')'
        curs.execute(query)
    query = 'SELECT {run} FROM {tn}'.format(tn=individual_list, run=sql_run)
    this_run = len(curs.execute(query).fetchall()) + 1
    query = 'INSERT INTO {tn} ({run}) VALUES (?)'.format(tn=individual_list, run=sql_run)
    curs.execute(query, (this_run,))
    for dancer in all_dancers:
        dancer_id = dancer[gen_dict['id']]
        query = 'SELECT * FROM {tn} WHERE {id} = ?'.format(id=sql_dict['id'], tn=selected_list)
        dancer_selected = len(curs.execute(query, (dancer_id,)).fetchall())
        query = 'UPDATE {tn} SET "{col}" = ? WHERE {run} = ?'.format(tn=individual_list, run=sql_run, col=dancer_id)
        curs.execute(query, (dancer_selected, this_run))
    conn.commit()
    # Update status
    status_update()
    # Close cursor and connection
    curs.close()
    conn.close()
    # Enable the CLI
    cli_text.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    root.geometry("1600x900")
    root.state('zoomed')
    root.title('NTDS 2018 Selection (BETA)')
    pad_out = 8
    pad_in = 8
    frame = Frame()
    frame.place(in_=root, anchor="c", relx=.50, rely=.50)
    x_scrollbar = Scrollbar(master=frame, orient=HORIZONTAL)
    x_scrollbar.grid(row=1, column=0, padx=pad_in, sticky=E+W)
    y_scrollbar = Scrollbar(master=frame, orient=VERTICAL)
    y_scrollbar.grid(row=0, column=1, pady=pad_in, sticky=N+S)
    status_text = Text(master=frame, width=status_text_width, height=50, padx=pad_in, pady=pad_in,
                       xscrollcommand=x_scrollbar.set, yscrollcommand=y_scrollbar.set, state=DISABLED, wrap=NONE)
    status_text.grid(row=0, column=0, padx=pad_out)
    x_scrollbar.config(command=status_text.xview)
    y_scrollbar.config(command=status_text.yview)
    cli_text = Entry(master=frame, width=int(status_text_width*27/20))
    cli_text.grid(row=2, column=0, padx=pad_in, pady=pad_out)
    cli_text.bind('<Return> ', cli_parser)
    data_help_frame = Frame(master=frame)
    data_help_frame.grid(row=0, column=2, rowspan=4, columnspan=3)
    data_text = Text(master=data_help_frame, width=total_width-status_text_width, height=50,
                     padx=pad_in, pady=pad_in, wrap=WORD, state=DISABLED)
    data_text.grid(row=0, column=0, padx=pad_out, columnspan=3)
    padding_frame = Frame(master=data_help_frame, height=16)
    padding_frame.grid(row=1, column=0)
    start_button = Button(master=data_help_frame, text='Start new selection database', command=main_selection)
    start_button.grid(row=2, column=0, padx=pad_out, pady=pad_in)
    update_button = Button(master=data_help_frame, text='Select existing database', command=select_database)
    update_button.grid(row=2, column=1, padx=pad_out)
    options_button = Button(master=data_help_frame, text='Settings', command=options_menu)
    options_button.grid(row=2, column=2, padx=pad_out)
    welcome_text()
    print_ntds_config()
    cli_text.focus_set()
    select_db = EntryBox
    select_db.root = root
    root.mainloop()
